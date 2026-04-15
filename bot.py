"""
╔══════════════════════════════════════════════╗
║   QARZ HISOBCHI BOT  v5.1  🏦               ║
║   Guruh + Shaxsiy | Smart Parser            ║
║   Ko'p valyuta | PDF | Excel | Undo         ║
║   TUZATILGAN: Kirilcha қарз, Excel, PDF     ║
╚══════════════════════════════════════════════╝
"""
import logging, re, os, random, io, tempfile, json
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer
from datetime import datetime, time as timeofday
from dotenv import load_dotenv
load_dotenv()

from telegram import (
    Update, ReplyKeyboardMarkup, KeyboardButton,
    InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardRemove, BotCommand, WebAppInfo,
    MenuButtonWebApp, MenuButtonCommands,
)
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    filters, ContextTypes, CallbackQueryHandler,
    ConversationHandler
)
from telegram.constants import ChatType

from database import Database
from reporter import (generate_report, generate_excel,
                      generate_pdf, generate_monthly_chart)
from lang import T, LANGS, CURRENCY_SYMBOLS

# ──────────────────────────────────────────────────────────────────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN", "")
if not BOT_TOKEN:
    raise ValueError("❌ BOT_TOKEN topilmadi! .env faylini tekshiring.")

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("bot.log", encoding="utf-8"),
    ]
)
logger = logging.getLogger(__name__)
# Prevent access token leakage in request logs (httpx prints full URLs on INFO).
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)
db = Database()

ADMIN_IDS = [int(x.strip()) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()]
_disable_webhook = os.getenv("DISABLE_WEBHOOK", "0").strip().lower() in ("1", "true", "yes", "on")
WEBHOOK_URL = "" if _disable_webhook else os.getenv("WEBHOOK_URL", os.getenv("RENDER_EXTERNAL_URL", "")).strip()
WEBHOOK_PORT = int(os.getenv("PORT", os.getenv("WEBHOOK_PORT", "8443")))
WEBHOOK_PATH = os.getenv("WEBHOOK_PATH", "/tg-webhook").strip().strip("/") or "tg-webhook"
MINI_APP_URL = os.getenv("MINI_APP_URL", "").strip()


class _HealthHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        self.send_response(200)
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.end_headers()
        self.wfile.write(b"ok")

    def log_message(self, format, *args):
        # Health checks are frequent; keep logs clean.
        return


def _start_health_server_if_needed():
    """
    Render Web Service expects an open port.
    When webhook is disabled (polling mode), start a tiny health server.
    """
    if not _disable_webhook:
        return
    port_raw = os.getenv("PORT", "").strip()
    if not port_raw.isdigit():
        return
    port = int(port_raw)

    def _run():
        try:
            server = HTTPServer(("0.0.0.0", port), _HealthHandler)
            logger.info("Health server started on port %s (polling mode)", port)
            server.serve_forever()
        except Exception as e:
            logger.warning("Health server start failed: %s", e)

    t = threading.Thread(target=_run, daemon=True)
    t.start()


def L(update: Update, key: str, **kwargs) -> str:
    if not update.effective_chat:
        return T(0, key, db, **kwargs)
    return T(update.effective_chat.id, key, db, **kwargs)


def _reply_msg(update: Update):
    if update.message:
        return update.message
    if update.callback_query and update.callback_query.message:
        return update.callback_query.message
    return update.effective_message


def _sum_by_currency(debts: list, field: str = 'remaining') -> dict[str, float]:
    out: dict[str, float] = {}
    for d in debts:
        cur = d.get('currency') or 'UZS'
        out[cur] = out.get(cur, 0) + d[field]
    return out


def _list_remaining_header(debts: list) -> str:
    by_r = _sum_by_currency(debts, 'remaining')
    if not by_r:
        return ""
    if len(by_r) == 1:
        cur, v = next(iter(by_r.items()))
        sym = CURRENCY_SYMBOLS.get(cur, cur)
        return f"💰 <b>{int(v):,}</b> {sym} qolgan\n".replace(',', ' ')
    parts = []
    for cur, v in sorted(by_r.items()):
        sym = CURRENCY_SYMBOLS.get(cur, cur)
        parts.append(f"<b>{int(v):,}</b> {sym}".replace(',', ' '))
    return f"💰 Qoldiq: {' · '.join(parts)}\n"


def _rates_approx_footer(debts: list, cid: int) -> str:
    usd = os.getenv("RATE_USD_UZS", "").strip()
    rub = os.getenv("RATE_RUB_UZS", "").strip()
    if not usd and not rub:
        return ""
    try:
        ru = float(usd) if usd else None
        rr = float(rub) if rub else None
    except ValueError:
        return ""
    by_r = _sum_by_currency(debts, 'remaining')
    lines = []
    if ru:
        u = by_r.get('USD', 0)
        if u:
            lines.append(f"USD→UZS ~ <b>{int(u * ru):,}</b>".replace(',', ' '))
    if rr:
        r = by_r.get('RUB', 0)
        if r:
            lines.append(f"RUB→UZS ~ <b>{int(r * rr):,}</b>".replace(',', ' '))
    if not lines:
        return ""
    return "\n" + "\n".join(lines) + T(cid, 'rate_footer', db)

# ──────────────────────────────────────────────────────────────────────────────
#  CONVERSATION STATES
# ──────────────────────────────────────────────────────────────────────────────
(S_NAME, S_CURRENCY, S_AMOUNT, S_CATEGORY,
 S_DUE, S_NOTE) = range(6)
WAIT_PAY = "wait_pay"

# ──────────────────────────────────────────────────────────────────────────────
#  KALIT SO'ZLAR — TUZATILGAN (kirilcha to'liq qo'llab-quvvatlash)
# ──────────────────────────────────────────────────────────────────────────────
#
# MUHIM: қарз so'zi kirilchada. Python string da saqlanganda UTF-8 bilan
# to'g'ri saqlanadi. re.search() da UNICODE flag avtomatik.
#
DEBT_KW = [
    'qarz', 'karz', 'qarzi', 'karzi',
    'қарз', 'қарзи', 'карз', 'карзи',       # kirilcha
    'oldi', 'qarz berdi', 'berdi qarz',
    'karzga', 'qarzga',
]
PAY_KW = [
    "berdi", "yopdi", "to'ladi", "toladi",
    "tashladi", "qaytardi", "tashlab ketdi",
    "uzdi", "uzib berdi", "qaytdi", "to'lab berdi",
    "берди", "тўлади", "қайтарди",           # kirilcha to'lov
]

NAME_IGNORE = set(
    DEBT_KW + PAY_KW +
    ["so'm", "som", "sum", "ming", "mln", "million",
     "dollar", "rubl", "usd", "uzs", "rub",
     "bu", "u", "men", "sen", "biz", "siz",
     "ha", "yo", "va", "ham", "yana", "lekin",
     "bor", "yo'q"]
)

# ──────────────────────────────────────────────────────────────────────────────
#  QUVNOQ JAVOBLAR
# ──────────────────────────────────────────────────────────────────────────────
FUN_DEBT = [
    "📝 Yozib qo'ydim! Hisobchi hech narsani unutmaydi 🧠",
    "✅ Bazaga qo'shildi! Qochib qutula olmaydi 😄",
    "💾 Saqlandi! Pul — hisob bilan 💪",
    "🎯 Bot kuzatib turadi! 24/7 uyg'oq 👀",
    "📌 Belgilab oldim! Hisob-kitob aniq bo'lsin 😎",
    "🔐 Qulflab qo'ydim! 🔒",
]
FUN_PAY = [
    "💳 To'lov qabul! Barakalla 👏",
    "✅ Vijdonli odam ekan! 😊",
    "💰 Hisobga olindi! Rahmat, halol odam 🤝",
    "📥 Qabul qilindi! Davom etaversin 🙌",
    "🎉 Pul qaytdi! To'lovga shay odam 🥳",
]
FUN_FULL = [
    "🎊 BARAKALLA! Qarz to'liq uzildi! 🎉✨",
    "🏆 CHAMPION! Qarz nolga tushdi! 🥳🎊",
    "🎉 YASHANG! Vijdon toza! To'liq to'landi! 🌟",
    "🎊 WOW! Qarz tugadi! Endi erkin odam! 🦅🎉",
    "🏅 SUPER! Zo'r odam! Qarz yo'q! 🎊🎈",
]

# ──────────────────────────────────────────────────────────────────────────────
#  KLAVIATURALAR
# ──────────────────────────────────────────────────────────────────────────────
MAIN_KB = ReplyKeyboardMarkup([
    [KeyboardButton("➕ Qarz qo'shish"),   KeyboardButton("💳 To'lov kiritish")],
    [KeyboardButton("📋 Ro'yxat"),          KeyboardButton("📊 Statistika")],
    [KeyboardButton("📈 Hisobot"),          KeyboardButton("📥 Excel")],
    [KeyboardButton("🔍 Qidirish"),         KeyboardButton("❓ Yordam")],
], resize_keyboard=True)

CANCEL_KB = ReplyKeyboardMarkup(
    [[KeyboardButton("🔙 Bekor qilish")]],
    resize_keyboard=True, one_time_keyboard=True
)
SKIP_CANCEL_KB = ReplyKeyboardMarkup(
    [[KeyboardButton("⏩ O'tkazib yuborish"), KeyboardButton("🔙 Bekor qilish")]],
    resize_keyboard=True, one_time_keyboard=True
)

def is_group(update: Update) -> bool:
    return update.effective_chat.type in [ChatType.GROUP, ChatType.SUPERGROUP]

# ──────────────────────────────────────────────────────────────────────────────
#  SMART PARSER — TUZATILGAN
#  Kirilcha қарз so'zini to'g'ri aniqlash
# ──────────────────────────────────────────────────────────────────────────────
def _contains_kw(text_lower: str, keywords: list) -> bool:
    """
    Kalit so'zni tekshiradi.
    Kirilcha so'zlar uchun ham to'g'ri ishlaydi.
    """
    for kw in keywords:
        if kw in text_lower:
            return True
    return False


def parse_debt(text: str) -> dict | None:
    """
    Qarz/to'lov matnini tahlil qiladi.
    Qo'llab-quvvatlanadigan formatlar:
      Alisher 500000 qarz
      Nodira karz 1.5 mln
      Bobur 300000 berdi
      @sardor 200000 to'ladi
      Inomjon 6000          (kalit so'zsiz)
      inomjon6000           (biriktirilgan)
      Ali 1.5mln
      inomjon 100000қарз    (kirilcha biriktirilgan) ← YANGI
      Nodira 150000 қарз    (kirilcha ajratilgan)     ← YANGI
    """
    if not text or len(text) > 500:
        return None

    tl = text.lower().strip()
    if len(tl) < 3:
        return None

    # ── To'lov va qarz kalit so'zlarini aniqlash ─────────────────────────────
    is_pay     = _contains_kw(tl, PAY_KW)
    is_debt_kw = _contains_kw(tl, DEBT_KW)

    # ── Summa aniqlash ────────────────────────────────────────────────────────
    amount = None

    # mln / million format
    m_mln = re.search(
        r'(\d[\d\s,\.]*)[\s]?(?:mln|million|mlrd)\b', text, re.IGNORECASE
    )
    m_k = re.search(r'(\d[\d\s,\.]*)[\s]?[kK]\b', text)

    if m_mln:
        try:
            num_str = m_mln.group(1).replace(',', '').replace(' ', '')
            # "1.5" ni float ga o'girish
            if num_str.count('.') <= 1:
                amount = int(float(num_str) * 1_000_000)
        except Exception:
            pass
    elif m_k:
        try:
            num_str = m_k.group(1).replace(',', '').replace(' ', '')
            amount = int(float(num_str) * 1_000)
        except Exception:
            pass

    if not amount:
        # Oddiy raqamlar: "6000", "6 000", "6,000"
        # Kirilcha bilan biriktirilgan ham: "100000қарз" → 100000
        # Raqamni kirilcha/lotin harflardan ajratib olamiz
        clean_text = text.replace('\u00a0', ' ')
        # Raqam + ixtiyoriy harflar (масалан 100000қарз)
        nums = re.findall(r'(\d[\d\s,]*\d|\d+)', clean_text)
        for n in nums:
            v = int(n.replace(',', '').replace(' ', ''))
            if v >= 100:
                amount = v
                break

    if not amount:
        return None

    # ── Ism aniqlash ──────────────────────────────────────────────────────────
    name = None

    # @username
    um = re.search(r'@(\w{3,32})', text)
    if um:
        name = "@" + um.group(1)

    if not name:
        # Lotin va krill harfli ismlar (3-25 belgi)
        words = re.findall(
            r'\b([A-ZА-ЯЎҚҒҲЁa-zA-Zа-яёўқғҳ]{3,25})\b', text
        )
        for w in words:
            wl = w.lower()
            if wl not in NAME_IGNORE:
                name = w
                break

    if not name:
        # "inomjon6000" yoki "inomjon100000қарз" — biriktirilgan
        m_glued = re.match(
            r'^([A-ZА-ЯЎҚҒҲЁa-zA-Zа-яёўқғҳ]{3,25})(\d+)', text.strip()
        )
        if m_glued:
            name   = m_glued.group(1)
            amount = int(m_glued.group(2))

    if not name:
        return None

    # ── Kalit so'zsiz qarz logikasi ───────────────────────────────────────────
    if not is_debt_kw and not is_pay:
        word_count = len(re.findall(r'\b\w{3,}\b', text))
        if word_count > 4:
            return None
        # Qisqa format: "Ism raqam" = qarz sifatida qabul qilish
        is_debt_kw = True

    if not is_debt_kw and not is_pay:
        return None

    return {
        'name':         name,
        'amount':       amount,
        'is_repayment': is_pay
    }


# ──────────────────────────────────────────────────────────────────────────────
#  QARZ KARTOCHKASI
# ──────────────────────────────────────────────────────────────────────────────
def debt_card(d: dict, idx: int = None) -> str:
    paid   = d['amount'] - d['remaining']
    pct    = int(paid / d['amount'] * 100) if d['amount'] > 0 else 0
    filled = int(10 * pct / 100)
    bar    = "🟩" * filled + "⬜" * (10 - filled)

    if   pct == 0:  badge = "🔴 Yangi"
    elif pct < 50:  badge = "🟡 Jarayonda"
    elif pct < 100: badge = "🟠 Yaqin"
    else:           badge = "✅ Tugadi"

    pfx  = f"{idx}. " if idx else ""
    dt   = (d.get('created_at') or '')[:10]
    cur  = d.get('currency', 'UZS')
    cs   = CURRENCY_SYMBOLS.get(cur, cur)
    cat  = f" · 🗂 <i>{d['category']}</i>" if d.get('category') else ""
    today = datetime.now().strftime('%Y-%m-%d')
    due_str = ""
    if d.get('due_date'):
        overdue = d['due_date'] < today and d['remaining'] > 0
        icon = "⚠️" if overdue else "📅"
        due_str = f"\n│ {icon} Muddat: <b>{d['due_date']}</b>"
        if overdue:
            due_str += " — <b>MUDDATI O'TDI!</b>"

    return (
        f"┌──────────────────────────\n"
        f"│ 👤 <b>{pfx}{d['name']}</b>  {badge}{cat}\n"
        f"│ 💸 Qarz:     <b>{int(d['amount']):,} {cs}</b>\n"
        f"│ ✔️ To'landi:  <b>{int(paid):,} {cs}</b>\n"
        f"│ ⏳ Qoldi:    <b>{int(d['remaining']):,} {cs}</b>\n"
        f"│ {bar} {pct}%{due_str}\n"
        f"│ 📅 {dt}  ·  🆔 #{d['id']}\n"
        f"└──────────────────────────"
    ).replace(',', ' ')


# ──────────────────────────────────────────────────────────────────────────────
#  RO'YXAT (sahifalash)
# ──────────────────────────────────────────────────────────────────────────────
PAGE = 3

async def send_list(update: Update, page: int = 0, edit: bool = False):
    cid   = update.effective_chat.id
    debts = db.get_active_debts(cid)

    if not debts:
        txt = (
            f"{L(update, 'no_debts')}\n\n"
            f"{L(update, 'list_empty_hint')}\n"
            "• <code>Alisher 500000 qarz</code>"
        )
        if edit:
            await update.callback_query.message.edit_text(
                txt, parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("➕ Qarz qo'shish", callback_data="how_debt")
                ]])
            )
        else:
            await _reply_msg(update).reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)
        return

    total_p = (len(debts) + PAGE - 1) // PAGE
    page    = max(0, min(page, total_p - 1))
    chunk   = debts[page * PAGE: (page + 1) * PAGE]

    hdr = (
        f"📋 <b>Qarzdorlar ro'yxati</b>\n"
        f"👥 <b>{len(debts)}</b> kishi  ·  "
        f"{_list_remaining_header(debts)}"
        f"━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    ).replace(',', ' ')

    body = "\n\n".join(
        debt_card(d, page * PAGE + i + 1)
        for i, d in enumerate(chunk)
    )
    txt = hdr + body

    btns = []
    for d in chunk:
        cs = CURRENCY_SYMBOLS.get(d.get('currency', 'UZS'), '')
        btns.append([
            InlineKeyboardButton(f"💳 {d['name'][:13]}", callback_data=f"pay_{d['id']}"),
            InlineKeyboardButton("📋", callback_data=f"hist_{d['id']}"),
            InlineKeyboardButton("🗑",  callback_data=f"del_{d['id']}"),
        ])

    if total_p > 1:
        nav = []
        if page > 0:
            nav.append(InlineKeyboardButton("⬅️", callback_data=f"pg_{page-1}"))
        nav.append(InlineKeyboardButton(f"📄 {page+1}/{total_p}", callback_data="noop"))
        if page < total_p - 1:
            nav.append(InlineKeyboardButton("➡️", callback_data=f"pg_{page+1}"))
        btns.append(nav)

    btns.append([
        InlineKeyboardButton("🔄 Yangilash",  callback_data=f"pg_{page}"),
        InlineKeyboardButton("📊 Statistika", callback_data="stats_inline"),
    ])
    kb = InlineKeyboardMarkup(btns)

    if edit:
        await update.callback_query.message.edit_text(txt, parse_mode='HTML', reply_markup=kb)
    else:
        await _reply_msg(update).reply_text(txt, parse_mode='HTML', reply_markup=kb)


# ──────────────────────────────────────────────────────────────────────────────
#  QARZ / TO'LOV QAYTA ISHLASH
# ──────────────────────────────────────────────────────────────────────────────
async def do_process(update: Update, text: str, cid: int,
                     is_group_chat: bool = False):
    info = parse_debt(text)
    if not info:
        return False

    u      = update.effective_user
    sender = (f"@{u.username}" if u and u.username
              else (u.full_name if u else "unknown"))
    uid    = u.id if u else 0
    mid    = 0
    if update.message and update.message.reply_to_message:
        mid = update.message.reply_to_message.message_id

    debt_id, new_rem, updated, meta = db.process_transaction(
        cid, info['name'], info['amount'],
        info['is_repayment'], added_by=sender,
        message_id=mid, user_id=uid,
    )

    af = f"{info['amount']:,}".replace(',', ' ')
    rf = f"{new_rem:,}".replace(',', ' ')
    drow = db.get_debt_by_id(debt_id, cid) if debt_id else None
    cur_code = (drow.get('currency') if drow else None) or 'UZS'
    cs = CURRENCY_SYMBOLS.get(cur_code, cur_code)

    if debt_id is None:
        await update.message.reply_text(
            f"❓ <b>{info['name']}</b> uchun faol qarz topilmadi!\n"
            "Avval qarz kiritilishi kerak.",
            parse_mode='HTML', reply_markup=MAIN_KB
        )
        return True

    if info['is_repayment']:
        if new_rem <= 0:
            msg = (
                f"{random.choice(FUN_FULL)}\n\n"
                f"👤 <b>{info['name']}</b>\n"
                f"💰 To'landi: <b>{af} {cs}</b>\n"
                f"🏁 Qarz yo'q! Jami uzildi 🎊"
            )
        else:
            msg = (
                f"{random.choice(FUN_PAY)}\n\n"
                f"👤 <b>{info['name']}</b>\n"
                f"💰 To'landi: <b>{af} {cs}</b>\n"
                f"📉 Hali qoldi: <b>{rf} {cs}</b>"
            )
        if meta and meta.get('repayment') and float(meta.get('benefit', 0)) > float(info['amount']) + 1e-6:
            pct = db.get_setting(cid, 'early_bonus_pct', '0')
            eff = f"{int(meta['benefit']):,}".replace(',', ' ')
            msg += "\n" + L(update, 'early_bonus_line', pct=pct, eff=eff)
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=MAIN_KB)
    else:
        act = "mavjud qarzga qo'shildi ➕" if updated else "yangi qarz 🆕"
        msg = (
            f"{random.choice(FUN_DEBT)}\n\n"
            f"┌──────────────────────────\n"
            f"│ 👤 <b>{info['name']}</b>\n"
            f"│ 💸 <b>{af} {cs}</b>\n"
            f"│ 📈 Jami qarzi: <b>{rf} {cs}</b>\n"
            f"│ 📌 {act}  ·  🆔 #{debt_id}\n"
            f"└──────────────────────────"
        )
        if mid:
            msg += "\n" + L(update, 'reply_linked')
        await update.message.reply_text(msg, parse_mode='HTML', reply_markup=MAIN_KB)
        await update.message.reply_text(
            "⚡ Tezkor amallar:",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("💳 To'lov kiritish", callback_data=f"pay_{debt_id}"),
                InlineKeyboardButton("↩️ Bekor qilish",   callback_data="undo_last"),
            ]])
        )
    return True


# ══════════════════════════════════════════════════════════════════════════════
#  /START
# ══════════════════════════════════════════════════════════════════════════════
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    name = user.first_name if user else "Do'stim"
    chat = update.effective_chat

    if is_group(update):
        await update.message.reply_text(
            f"👋 <b>Salom, {chat.title or 'Guruh'}!</b>\n\n"
            "🏦 <b>Qarz Hisobchi Bot v5.1</b> faol!\n\n"
            "📝 <b>Qarz kiritish:</b>\n"
            "  <code>Alisher 500000 qarz</code>\n"
            "  <code>Nodira 1.5 mln karz</code>\n"
            "  <code>Inomjon 6000</code>  ← kalit so'zsiz!\n"
            "  <code>Ali 150000қарз</code> ← kirilcha!\n\n"
            "💳 <b>To'lov:</b>\n"
            "  <code>Alisher 200000 berdi</code>\n\n"
            "📊 /qarzlar · /top · /hisobot · /yordam",
            parse_mode='HTML', reply_markup=MAIN_KB
        )
        return

    await update.message.reply_text(
        f"🏦 <b>QARZ HISOBCHI BOT v5.1</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"👋 Salom, <b>{name}</b>!\n\n"
        f"📥 <b>Qarz kiritish:</b>\n"
        f"  <code>Alisher 500 000 qarz</code>\n"
        f"  <code>Nodira 1.5 mln karz</code>\n"
        f"  <code>@sardor 300000 қарз</code>\n"
        f"  <code>Inomjon 6000</code>  ← kalit so'zsiz!\n"
        f"  <code>inomjon6000</code>   ← biriktirilgan!\n"
        f"  <code>Ali 100000қарз</code> ← kirilcha!\n\n"
        f"💳 <b>To'lov kiritish:</b>\n"
        f"  <code>Alisher 200 000 berdi</code>\n"
        f"  <code>Nodira to'ladi 500 000</code>\n"
        f"  <code>Bobur qaytardi 100 000</code>\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"👇 <b>Pastdagi tugmalardan foydalaning!</b>",
        parse_mode='HTML', reply_markup=MAIN_KB
    )


# ══════════════════════════════════════════════════════════════════════════════
#  QARZ QO'SHISH — ConversationHandler
# ══════════════════════════════════════════════════════════════════════════════
async def wizard_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🪄 <b>Qarz qo'shish — 1/6</b>\n\n" + L(update, 'enter_name') + "\n"
        "<i>Misol: Alisher  yoki  @username</i>",
        parse_mode='HTML', reply_markup=CANCEL_KB
    )
    return S_NAME

async def w_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tx = update.message.text.strip()
    if tx == "🔙 Bekor qilish":
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        return ConversationHandler.END
    context.user_data['dn'] = tx

    kb = ReplyKeyboardMarkup([
        [KeyboardButton("🇺🇿 So'm (UZS)"),
         KeyboardButton("🇺🇸 Dollar (USD)"),
         KeyboardButton("🇷🇺 Rubl (RUB)")],
        [KeyboardButton("🔙 Bekor qilish")],
    ], resize_keyboard=True, one_time_keyboard=True)

    await update.message.reply_text(
        f"✅ Ism: <b>{tx}</b>\n\n"
        "🪄 <b>Qarz qo'shish — 2/6</b>\n\n"
        "💱 <b>Valyutani</b> tanlang:",
        parse_mode='HTML', reply_markup=kb
    )
    return S_CURRENCY

async def w_currency(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tx = update.message.text.strip()
    if tx == "🔙 Bekor qilish":
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        return ConversationHandler.END
    cur_map = {
        "🇺🇿 So'm (UZS)": "UZS",
        "🇺🇸 Dollar (USD)": "USD",
        "🇷🇺 Rubl (RUB)": "RUB",
        "UZS": "UZS", "USD": "USD", "RUB": "RUB",
    }
    context.user_data['dc'] = cur_map.get(tx, 'UZS')
    await update.message.reply_text(
        f"✅ Valyuta: <b>{context.user_data['dc']}</b>\n\n"
        "🪄 <b>Qarz qo'shish — 3/6</b>\n\n"
        "💰 <b>Summasini</b> yozing:\n"
        "<i>Misol: 500000  yoki  1.5 mln</i>",
        parse_mode='HTML', reply_markup=CANCEL_KB
    )
    return S_AMOUNT

async def w_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tx = update.message.text.strip()
    if tx == "🔙 Bekor qilish":
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        return ConversationHandler.END

    amount = _parse_amount(tx)
    if not amount:
        await update.message.reply_text(
            "❌ <b>Noto'g'ri summa!</b>\n\n"
            "Misol: <code>500000</code>  yoki  <code>1.5 mln</code>",
            parse_mode='HTML', reply_markup=CANCEL_KB
        )
        return S_AMOUNT

    context.user_data['da'] = amount
    cs = CURRENCY_SYMBOLS.get(context.user_data.get('dc', 'UZS'), "so'm")

    cat_kb = ReplyKeyboardMarkup([
        [KeyboardButton("🛒 Mahsulot"), KeyboardButton("🔧 Xizmat")],
        [KeyboardButton("👤 Shaxsiy"),  KeyboardButton("📦 Boshqa")],
        [KeyboardButton("⏩ O'tkazib yuborish"), KeyboardButton("🔙 Bekor qilish")],
    ], resize_keyboard=True, one_time_keyboard=True)

    await update.message.reply_text(
        f"✅ Summa: <b>{amount:,} {cs}</b>\n\n".replace(',', ' ') +
        "🪄 <b>Qarz qo'shish — 4/6</b>\n\n"
        "🗂 <b>Kategoriyani</b> tanlang:",
        parse_mode='HTML', reply_markup=cat_kb
    )
    return S_CATEGORY

async def w_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tx = update.message.text.strip()
    if tx == "🔙 Bekor qilish":
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        return ConversationHandler.END
    cat_map = {
        "🛒 Mahsulot": "Mahsulot",
        "🔧 Xizmat":   "Xizmat",
        "👤 Shaxsiy":  "Shaxsiy",
        "📦 Boshqa":   "Boshqa",
    }
    category = "" if tx == "⏩ O'tkazib yuborish" else cat_map.get(tx, tx)
    context.user_data['dcat'] = category

    await update.message.reply_text(
        (f"✅ Kategoriya: <b>{category}</b>\n\n" if category else "") +
        "🪄 <b>Qarz qo'shish — 5/6</b>\n\n"
        "📅 <b>Muddat</b> kiriting (ixtiyoriy):\n"
        "<i>Format: 25.12.2025</i>",
        parse_mode='HTML', reply_markup=SKIP_CANCEL_KB
    )
    return S_DUE

async def w_due(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tx = update.message.text.strip()
    if tx == "🔙 Bekor qilish":
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        return ConversationHandler.END
    due_date = None
    if tx not in ("⏩ O'tkazib yuborish", ""):
        try:
            d = datetime.strptime(tx, "%d.%m.%Y")
            due_date = d.strftime("%Y-%m-%d")
        except ValueError:
            await update.message.reply_text(
                "❌ Format noto'g'ri!\n<i>To'g'ri: 25.12.2025</i>",
                parse_mode='HTML', reply_markup=SKIP_CANCEL_KB
            )
            return S_DUE
    context.user_data['ddue'] = due_date
    await update.message.reply_text(
        "🪄 <b>Qarz qo'shish — 6/6</b>\n\n"
        "📝 <b>Izoh</b> yozing (ixtiyoriy):\n"
        "<i>Misol: Oylik, Do'kondan, ...</i>",
        parse_mode='HTML', reply_markup=SKIP_CANCEL_KB
    )
    return S_NOTE

async def w_note(update: Update, context: ContextTypes.DEFAULT_TYPE):
    tx = update.message.text.strip()
    if tx == "🔙 Bekor qilish":
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        context.user_data.clear()
        return ConversationHandler.END

    note     = "" if tx == "⏩ O'tkazib yuborish" else tx
    name     = context.user_data.get('dn', '')
    amount   = context.user_data.get('da', 0)
    currency = context.user_data.get('dc', 'UZS')
    category = context.user_data.get('dcat', '')
    due_date = context.user_data.get('ddue', None)
    u        = update.effective_user
    sender   = f"@{u.username}" if u.username else u.full_name
    uid      = u.id if u else 0
    mid      = 0
    if update.message.reply_to_message:
        mid = update.message.reply_to_message.message_id

    debt_id, new_rem, updated, _meta = db.process_transaction(
        update.effective_chat.id, name, amount,
        False, note, sender, mid, due_date, category, currency,
        user_id=uid,
    )

    act = "mavjud qarzga qo'shildi ➕" if updated else "yangi qarz yaratildi 🆕"
    cs  = CURRENCY_SYMBOLS.get(currency, "so'm")
    msg = (
        f"{random.choice(FUN_DEBT)}\n\n"
        f"┌──────────────────────────\n"
        f"│ 👤 <b>{name}</b>\n"
        f"│ 💸 <b>{amount:,} {cs}</b>\n".replace(',', ' ')
    )
    if category: msg += f"│ 🗂 {category}\n"
    if due_date: msg += f"│ 📅 Muddat: {due_date}\n"
    if note:     msg += f"│ 📝 {note}\n"
    msg += (
        f"│ 📌 {act}\n"
        f"│ 🆔 #{debt_id}\n"
        f"└──────────────────────────"
    )
    await update.message.reply_text(msg, parse_mode='HTML', reply_markup=MAIN_KB)
    context.user_data.clear()
    return ConversationHandler.END

async def wizard_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
    context.user_data.clear()
    return ConversationHandler.END


def _parse_amount(tx: str) -> int | None:
    """Satrdan summani ajratib oladi"""
    m_mln = re.search(r'(\d[\d\s,\.]*)[\s]?(?:mln|million|mlrd)\b', tx, re.IGNORECASE)
    m_k   = re.search(r'(\d[\d\s,\.]*)[\s]?[kK]\b', tx)
    if m_mln:
        try:
            return int(float(m_mln.group(1).replace(',', '').replace(' ', '')) * 1_000_000)
        except Exception:
            pass
    if m_k:
        try:
            return int(float(m_k.group(1).replace(',', '').replace(' ', '')) * 1_000)
        except Exception:
            pass
    c = tx.replace(',', '').replace(' ', '').replace('.', '')
    if c.isdigit() and int(c) >= 1:
        return int(c)
    return None


# ══════════════════════════════════════════════════════════════════════════════
#  STATISTIKA
# ══════════════════════════════════════════════════════════════════════════════
async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid   = update.effective_chat.id
    debts = db.get_active_debts(cid)

    if not debts:
        await update.message.reply_text(
            L(update, 'no_debts'), parse_mode='HTML', reply_markup=MAIN_KB
        )
        return

    currencies = {d.get('currency', 'UZS') for d in debts}
    top    = max(debts, key=lambda x: x['remaining'])
    least  = min(debts, key=lambda x: x['remaining'])
    cats   = db.get_categories(cid)
    ov     = db.get_overdue_debts(cid)

    if len(currencies) == 1:
        cur = next(iter(currencies))
        sym = CURRENCY_SYMBOLS.get(cur, cur)
        orig   = sum(d['amount']    for d in debts)
        rem    = sum(d['remaining'] for d in debts)
        paid   = orig - rem
        pct    = int(paid / orig * 100) if orig > 0 else 0
        filled = int(20 * pct / 100)
        bar    = "🟩" * filled + "⬜" * (20 - filled)
        txt = (
            f"📊 <b>UMUMIY STATISTIKA</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👥 Qarzdorlar:    <b>{len(debts)} kishi</b>\n"
            f"💸 Umumiy qarz:   <b>{int(orig):,} {sym}</b>\n"
            f"✅ To'langan:     <b>{int(paid):,} {sym}</b>\n"
            f"🔴 Qolgan:        <b>{int(rem):,} {sym}</b>\n\n"
            f"📈 Jarayon:\n"
            f"  {bar} <b>{pct}%</b>\n\n"
            f"🏆 Eng ko'p: <b>{top['name']}</b>\n"
            f"   — {int(top['remaining']):,} {CURRENCY_SYMBOLS.get(top.get('currency', 'UZS'), '')}\n\n"
            f"⚡ Eng oz: <b>{least['name']}</b>\n"
            f"   — {int(least['remaining']):,} {CURRENCY_SYMBOLS.get(least.get('currency', 'UZS'), '')}\n"
        ).replace(',', ' ')
    else:
        txt = (
            f"📊 <b>UMUMIY STATISTIKA</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"👥 Qarzdorlar:    <b>{len(debts)} kishi</b>\n"
            f"💱 <b>Turli valyutalar</b>:\n\n"
        )
        by_orig = _sum_by_currency(debts, 'amount')
        by_rem = _sum_by_currency(debts, 'remaining')
        for cur in sorted(set(by_orig) | set(by_rem)):
            sym = CURRENCY_SYMBOLS.get(cur, cur)
            o, r = by_orig.get(cur, 0), by_rem.get(cur, 0)
            p = o - r
            pct = int(p / o * 100) if o > 0 else 0
            txt += (
                f"  <b>{sym}</b> umumiy: <b>{int(o):,}</b> · "
                f"to'langan: <b>{int(p):,}</b> · qoldi: <b>{int(r):,}</b> ({pct}%)\n"
            ).replace(',', ' ')
        txt += (
            f"\n🏆 Eng ko'p: <b>{top['name']}</b>\n"
            f"   — {int(top['remaining']):,} {CURRENCY_SYMBOLS.get(top.get('currency', 'UZS'), '')}\n\n"
            f"⚡ Eng oz: <b>{least['name']}</b>\n"
            f"   — {int(least['remaining']):,} {CURRENCY_SYMBOLS.get(least.get('currency', 'UZS'), '')}\n"
        ).replace(',', ' ')

    if cats:
        txt += "\n🗂 <b>Kategoriyalar:</b>\n"
        for c in cats[:4]:
            ccy = c.get('currency', 'UZS')
            csym = CURRENCY_SYMBOLS.get(ccy, ccy)
            cat_label = c['category']
            if len(currencies) > 1:
                cat_label = f"{cat_label} ({ccy})"
            txt += f"  · {cat_label}: <b>{int(c['total']):,}</b> {csym}\n".replace(',', ' ')

    if ov:
        txt += f"\n⚠️ <b>Muddati o'tgan: {len(ov)} ta!</b> /muddat\n"

    txt += _rates_approx_footer(debts, cid)
    txt += f"\n📅 <i>{datetime.now().strftime('%d.%m.%Y %H:%M')}</i>"

    await update.message.reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)
    await update.message.reply_text(
        "📊 Ko'proq:",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("🏆 TOP-5",         callback_data="top5"),
            InlineKeyboardButton("📅 Oylik grafik",  callback_data="monthly_chart"),
            InlineKeyboardButton("📈 Hisobot",       callback_data="full_report"),
        ]])
    )


# ══════════════════════════════════════════════════════════════════════════════
#  ASOSIY XABAR HANDLER
# ══════════════════════════════════════════════════════════════════════════════
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return

    text = update.message.text.strip()
    cid  = update.effective_chat.id
    grp  = is_group(update)

    # ── To'lov kutish holati ──────────────────────────────────────────────────
    if WAIT_PAY in context.user_data:
        state = context.user_data.pop(WAIT_PAY)
        if text in ("🔙 Bekor qilish", "❌ Bekor"):
            await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
            return
        try:
            amount = int(text.replace(' ', '').replace(',', ''))
            uid = update.effective_user.id if update.effective_user else 0
            result = db.add_payment(state['debt_id'], cid, amount, user_id=uid)
            if not result['success']:
                await update.message.reply_text(f"❌ {result['error']}", reply_markup=MAIN_KB)
                return
            nr = result['remaining']
            pcs = CURRENCY_SYMBOLS.get(result.get('currency', 'UZS'), "so'm")
            af = f"{amount:,}".replace(',', ' ')
            rf = f"{nr:,}".replace(',', ' ')
            if nr <= 0:
                msg = (f"{random.choice(FUN_FULL)}\n\n"
                       f"👤 <b>{result['name']}</b>\n"
                       f"💰 To'landi: <b>{af} {pcs}</b>\n"
                       f"🏁 Qarz yo'q! 🎊")
            else:
                msg = (f"{random.choice(FUN_PAY)}\n\n"
                       f"👤 <b>{result['name']}</b>\n"
                       f"💰 To'landi: <b>{af} {pcs}</b>\n"
                       f"📉 Qoldi: <b>{rf} {pcs}</b>")
            if result.get('bonus_applied'):
                eff = f"{int(result.get('benefit', amount)):,}".replace(',', ' ')
                pct = db.get_setting(cid, 'early_bonus_pct', '0')
                msg += "\n" + L(update, 'early_bonus_line', pct=pct, eff=eff)
            await update.message.reply_text(msg, parse_mode='HTML', reply_markup=MAIN_KB)
        except ValueError:
            await update.message.reply_text(
                "❌ <b>Faqat raqam kiriting!</b>\nMisol: <code>200000</code>",
                parse_mode='HTML', reply_markup=CANCEL_KB
            )
            context.user_data[WAIT_PAY] = state
        return

    # ── Qidirish holati ───────────────────────────────────────────────────────
    if context.user_data.get('searching'):
        context.user_data.pop('searching')
        all_d = db.get_active_debts(cid)
        res   = [d for d in all_d if text.lower() in d['name'].lower()]
        if not res:
            await update.message.reply_text(
                f"🔍 <b>«{text}»</b> — topilmadi 😕",
                parse_mode='HTML', reply_markup=MAIN_KB
            )
        else:
            cards = "\n\n".join(debt_card(d, i + 1) for i, d in enumerate(res))
            await update.message.reply_text(
                f"🔍 <b>«{text}»</b> — {len(res)} natija:\n\n{cards}",
                parse_mode='HTML', reply_markup=MAIN_KB
            )
        return

    # ── MENYU TUGMALARI ───────────────────────────────────────────────────────
    if text == "➕ Qarz qo'shish":
        await update.message.reply_text(
            "➕ <b>Qarz qo'shish</b>\n\nUsulni tanlang:",
            parse_mode='HTML', reply_markup=MAIN_KB
        )
        await update.message.reply_text(
            "👇", reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("📝 Tez yozish",        callback_data="how_debt"),
                InlineKeyboardButton("🪄 Bosqichma-bosqich", callback_data="wizard_debt"),
            ]])
        )
        return

    if text == "💳 To'lov kiritish":
        debts = db.get_active_debts(cid)
        if not debts:
            await update.message.reply_text("📭 <b>Faol qarzlar yo'q!</b>",
                                             parse_mode='HTML', reply_markup=MAIN_KB)
            return
        btns = []
        for _d in debts[:10]:
            _cs = CURRENCY_SYMBOLS.get(_d.get('currency', 'UZS'), "so'm")
            _lbl = f"💳 {_d['name'][:15]} — {int(_d['remaining']):,} {_cs}".replace(',', ' ')
            btns.append([InlineKeyboardButton(_lbl, callback_data=f"pay_{_d['id']}")])
        await update.message.reply_text(
            "💳 <b>Kim to'lov qildi?</b>", parse_mode='HTML', reply_markup=MAIN_KB
        )
        await update.message.reply_text("👇", reply_markup=InlineKeyboardMarkup(btns))
        return

    if text == "📋 Ro'yxat":
        await send_list(update, 0)
        return

    if text == "📊 Statistika":
        await cmd_stats(update, context)
        return

    if text == "📈 Hisobot":
        debts = db.get_active_debts(cid)
        if not debts:
            await update.message.reply_text("📭 Faol qarzlar yo'q!", reply_markup=MAIN_KB)
            return
        m = await update.message.reply_text(
            "⏳ <b>Hisobot tayyorlanmoqda...</b> 🔄", parse_mode='HTML'
        )
        try:
            res = generate_report(debts, cid)
            await m.delete()
            # Matn 4096 belgidan oshmasin
            report_text = res['text']
            if len(report_text) > 1020:
                report_text = report_text[:1020] + "\n\n_...qolganlar Excel/PDF da_"

            if res.get('chart_path') and os.path.exists(res['chart_path']):
                with open(res['chart_path'], 'rb') as f:
                    await update.message.reply_photo(
                        photo=f, caption=report_text,
                        parse_mode='Markdown', reply_markup=MAIN_KB
                    )
                os.remove(res['chart_path'])
            else:
                await update.message.reply_text(
                    report_text, parse_mode='Markdown', reply_markup=MAIN_KB
                )
        except Exception as e:
            logger.error(f"Hisobot xatosi: {e}")
            try:
                await m.delete()
            except Exception:
                pass
            await update.message.reply_text(
                f"❌ Hisobot xatosi: {e}", reply_markup=MAIN_KB
            )
        await update.message.reply_text(
            "📄 Boshqa formatlar:",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("📥 Excel",        callback_data="get_excel"),
                InlineKeyboardButton("📄 PDF",          callback_data="get_pdf"),
                InlineKeyboardButton("📅 Oylik grafik", callback_data="monthly_chart"),
            ]])
        )
        return

    if text == "📥 Excel":
        debts = db.get_active_debts(cid)
        if not debts:
            await update.message.reply_text("📭 Faol qarzlar yo'q!", reply_markup=MAIN_KB)
            return
        m = await update.message.reply_text("⏳ <b>Excel tayyorlanmoqda...</b> 📊", parse_mode='HTML')
        try:
            ep = generate_excel(debts, db.get_all_debts(cid), cid)
            await m.delete()
            with open(ep, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=f"qarz_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    caption=(
                        f"📥 <b>Excel tayyor!</b>\n"
                        f"📊 {len(debts)} ta qarzdor\n"
                        f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
                        f"<i>3 ta varaq: Faol qarzlar, Tarix, Statistika+Grafik</i>"
                    ),
                    parse_mode='HTML', reply_markup=MAIN_KB
                )
            os.remove(ep)
        except Exception as e:
            logger.error(f"Excel xatosi: {e}")
            try:
                await m.delete()
            except Exception:
                pass
            await update.message.reply_text(f"❌ Excel xatosi: {e}", reply_markup=MAIN_KB)
        return

    if text == "🔍 Qidirish":
        context.user_data['searching'] = True
        await update.message.reply_text(
            "🔍 <b>Qidirish</b>\n\nQarzdor <b>ismini</b> yozing:",
            parse_mode='HTML', reply_markup=CANCEL_KB
        )
        return

    if text == "❓ Yordam":
        await cmd_help(update, context)
        return

    if text == "🔙 Bekor qilish":
        context.user_data.clear()
        await update.message.reply_text("↩️ Bekor qilindi.", reply_markup=MAIN_KB)
        return

    # ── GURUH/SHAXSIY: qarz/to'lov matnini tahlil qilish ─────────────────────
    processed = await do_process(update, text, cid, grp)
    if not processed and not grp:
        # Mobile Telegram may hide keyboard; resend menu on unknown input.
        await update.message.reply_text(
            "ℹ️ Tushunmadim. Pastdagi menyudan tanlang yoki /menu yuboring.",
            reply_markup=MAIN_KB
        )


def _panel_text(chat_id: int) -> str:
    debts = db.get_active_debts(chat_id)
    if not debts:
        return T(chat_id, 'no_debts', db)
    n = len(debts)
    hdr = _list_remaining_header(debts).strip()
    top = max(debts, key=lambda x: x['remaining'])
    cs = CURRENCY_SYMBOLS.get(top.get('currency', 'UZS'), '')
    return (
        f"{T(chat_id, 'panel_title', db)}\n"
        f"━━━━━━━━━━━━━━\n"
        f"👥 <b>{n}</b>  ·  {hdr}\n"
        f"🏆 TOP: <b>{top['name']}</b> — <b>{int(top['remaining']):,}</b> {cs}".replace(',', ' ')
    )


def _panel_kb(chat_id: int) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton(T(chat_id, 'btn_refresh', db), callback_data="panel_refresh"),
            InlineKeyboardButton(T(chat_id, 'btn_list', db), callback_data="panel_list"),
        ],
        [
            InlineKeyboardButton(T(chat_id, 'btn_stats', db), callback_data="stats_inline"),
            InlineKeyboardButton(T(chat_id, 'btn_excel', db), callback_data="get_excel"),
            InlineKeyboardButton(T(chat_id, 'btn_chart', db), callback_data="monthly_chart"),
        ],
    ]
    if MINI_APP_URL:
        rows.append([InlineKeyboardButton(
            T(chat_id, 'miniapp_btn', db), web_app=WebAppInfo(url=MINI_APP_URL)
        )])
    return InlineKeyboardMarkup(rows)


async def cmd_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    await update.message.reply_text(
        _panel_text(cid) + "\n\n" + T(cid, 'panel_hint', db),
        parse_mode='HTML', reply_markup=_panel_kb(cid)
    )


async def cmd_audit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    lim = 15
    if context.args and context.args[0].isdigit():
        lim = min(50, max(1, int(context.args[0])))
    rows = db.get_audit(cid, lim)
    if not rows:
        await update.message.reply_text(L(update, 'audit_empty'), reply_markup=MAIN_KB)
        return
    lines = [L(update, 'audit_title')]
    for r in rows:
        try:
            d = json.loads(r['detail_json'] or '{}')
        except Exception:
            d = {}
        lines.append(
            f"• <code>{r['created_at'][:16]}</code> <b>{r['action']}</b> <code>{str(d)[:100]}</code>"
        )
    await update.message.reply_text("\n".join(lines)[:4000], parse_mode='HTML', reply_markup=MAIN_KB)


async def cmd_eslatmakun(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    if not context.args:
        days = int(db.get_setting(cid, 'reminder_days', '10') or 10)
        await update.message.reply_text(L(update, 'notify_cur', days=days), reply_markup=MAIN_KB)
        return
    try:
        d = int(context.args[0])
        if d < 1 or d > 365:
            raise ValueError
    except ValueError:
        await update.message.reply_text(L(update, 'notify_bad'), reply_markup=MAIN_KB)
        return
    db.set_setting(cid, 'reminder_days', str(d))
    await update.message.reply_text(L(update, 'notify_ok', days=d), reply_markup=MAIN_KB)


async def cmd_bonus(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    if not context.args:
        p = db.get_setting(cid, 'early_bonus_pct', '0')
        await update.message.reply_text(
            f"🎁 <b>Muddatdan oldin bonus</b>: <code>{p}%</code>\n"
            f"<i>/bonus 5</i> yoki <i>/bonus 0</i> — o'chirish",
            parse_mode='HTML', reply_markup=MAIN_KB,
        )
        return
    try:
        p = float(context.args[0].replace(',', '.'))
        if p < 0 or p > 50:
            raise ValueError
    except ValueError:
        await update.message.reply_text(L(update, 'bonus_bad'), reply_markup=MAIN_KB)
        return
    db.set_setting(cid, 'early_bonus_pct', str(p))
    await update.message.reply_text(L(update, 'bonus_ok', pct=p), reply_markup=MAIN_KB)


async def cmd_kurs(update: Update, context: ContextTypes.DEFAULT_TYPE):
    usd = os.getenv("RATE_USD_UZS", "").strip()
    rub = os.getenv("RATE_RUB_UZS", "").strip()
    if not usd and not rub:
        await update.message.reply_text(L(update, 'kurs_empty'), reply_markup=MAIN_KB)
        return
    txt = L(update, 'kurs_title') + "\n"
    if usd:
        txt += f"USD→UZS: <code>{usd}</code>\n"
    if rub:
        txt += f"RUB→UZS: <code>{rub}</code>\n"
    await update.message.reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)


# ══════════════════════════════════════════════════════════════════════════════
#  CALLBACK HANDLER
# ══════════════════════════════════════════════════════════════════════════════
async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q    = update.callback_query
    cid  = q.message.chat_id
    data = q.data

    if data.startswith("pg_"):
        await q.answer()
        page = int(data.split("_")[1])
        update._effective_chat = q.message.chat
        await send_list(update, page, edit=True)
        return

    if data == "noop":
        await q.answer()
        return

    if data == "stats_inline":
        debts = db.get_active_debts(cid)
        if not debts:
            await q.answer("📭 Qarz yo'q!", show_alert=True)
            return
        await q.answer()
        curs = {d.get('currency', 'UZS') for d in debts}
        if len(curs) == 1:
            orig = sum(d['amount'] for d in debts)
            rem = sum(d['remaining'] for d in debts)
            paid = orig - rem
            pct = int(paid / orig * 100) if orig > 0 else 0
            bar = "🟩" * int(20 * pct / 100) + "⬜" * (20 - int(20 * pct / 100))
            sym = CURRENCY_SYMBOLS.get(next(iter(curs)), "so'm")
            txt = (
                f"📊 <b>Statistika</b>\n"
                f"👥 {len(debts)} kishi · 💰 {int(rem):,} {sym}\n"
                f"{bar} {pct}%"
            ).replace(',', ' ')
        else:
            rem_line = _list_remaining_header(debts).strip()
            txt = (
                f"📊 <b>Statistika</b>\n"
                f"👥 {len(debts)} kishi · {rem_line}\n"
                f"<i>Turli valyuta</i>"
            )
        await q.message.reply_text(txt, parse_mode='HTML')
        return

    if data == "panel_refresh":
        await q.answer()
        await q.message.edit_text(_panel_text(cid), parse_mode='HTML', reply_markup=_panel_kb(cid))
        return

    if data == "panel_list":
        await q.answer()
        await send_list(update, 0, edit=False)
        return

    if data.startswith("pay_"):
        debt_id = int(data.split("_")[1])
        debt    = db.get_debt_by_id(debt_id, cid)
        if not debt:
            await q.answer("❌ Topilmadi!", show_alert=True)
            return
        await q.answer()
        context.user_data[WAIT_PAY] = {'debt_id': debt_id, 'name': debt['name']}
        cs = CURRENCY_SYMBOLS.get(debt.get('currency', 'UZS'), "so'm")
        await q.message.reply_text(
            f"💳 <b>{debt['name']}</b> uchun to'lov\n\n"
            f"📉 Qolgan qarz: <b>{int(debt['remaining']):,} {cs}</b>\n\n"
            "💰 <b>To'lov miqdorini kiriting:</b>".replace(',', ' '),
            parse_mode='HTML',
            reply_markup=ReplyKeyboardMarkup(
                [[KeyboardButton("🔙 Bekor qilish")]],
                resize_keyboard=True, one_time_keyboard=True
            )
        )
        return

    if data.startswith("del_"):
        debt = db.get_debt_by_id(int(data.split("_")[1]), cid)
        if not debt:
            await q.answer("❌ Topilmadi!", show_alert=True)
            return
        await q.answer()
        cs = CURRENCY_SYMBOLS.get(debt.get('currency', 'UZS'), "so'm")
        await q.message.reply_text(
            f"⚠️ <b>Tasdiqlang!</b>\n\n"
            f"👤 <b>{debt['name']}</b>\n"
            f"💰 Qolgan: <b>{int(debt['remaining']):,} {cs}</b>\n\n"
            "Bu qarzni o'chirmoqchimisiz?".replace(',', ' '),
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("✅ Ha, o'chir", callback_data=f"delok_{debt['id']}"),
                InlineKeyboardButton("❌ Bekor",      callback_data="cancel"),
            ]])
        )
        return

    if data.startswith("delok_"):
        await q.answer()
        debt_id = int(data.split("_")[1])
        debt    = db.get_debt_by_id(debt_id, cid)
        name    = debt['name'] if debt else f"#{debt_id}"
        db.delete_debt(debt_id, cid)
        await q.message.edit_text(
            f"🗑️ <b>{name}</b> ning qarz yozuvi o'chirildi. ✅",
            parse_mode='HTML'
        )
        return

    if data == "cancel":
        await q.answer()
        await q.message.edit_text("↩️ Bekor qilindi.")
        return

    if data.startswith("hist_"):
        debt_id    = int(data.split("_")[1])
        debt, pays = db.get_payment_history(debt_id, cid)
        if not debt:
            await q.answer("❌ Topilmadi!", show_alert=True)
            return
        await q.answer()
        cs  = CURRENCY_SYMBOLS.get(debt.get('currency', 'UZS'), "so'm")
        txt = (
            f"📋 <b>To'lov tarixi — {debt['name']}</b>\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"💸 Umumiy: <b>{int(debt['amount']):,} {cs}</b>\n"
            f"⏳ Qolgan: <b>{int(debt['remaining']):,} {cs}</b>\n\n"
        ).replace(',', ' ')
        if not pays:
            txt += "💳 Hali to'lovlar amalga oshirilmagan."
        else:
            for i, p in enumerate(pays, 1):
                txt += f"  {i}. {p['paid_at'][:10]} — <b>{int(p['amount']):,} {cs}</b>\n".replace(',', ' ')
            txt += f"\n✅ Jami: <b>{int(sum(p['amount'] for p in pays)):,} {cs}</b>".replace(',', ' ')
        await q.message.reply_text(txt, parse_mode='HTML')
        return

    if data == "undo_last":
        entry = db.pop_undo(cid)
        if not entry:
            await q.answer("❌ Bekor qilinadigan amal topilmadi!", show_alert=True)
            return
        await q.answer()
        name = entry['data'].get('name', '?')
        desc = {
            'ADD_DEBT':    "Qarz qo'shish",
            'PAYMENT':     "To'lov",
            'DELETE_DEBT': "O'chirish",
            'UPDATE_DEBT': "Yangilash",
        }.get(entry['action_type'], entry['action_type'])
        await q.message.reply_text(
            f"↩️ <b>Bekor qilindi!</b>\n\nAmal: {desc}\nKim: <b>{name}</b>",
            parse_mode='HTML'
        )
        return

    if data == "top5":
        await q.answer()
        top    = db.get_top_debtors(cid, 5)
        medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
        today  = datetime.now().strftime('%Y-%m-%d')
        txt    = "🏆 <b>TOP-5 QARZDORLAR</b>\n━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        for i, d in enumerate(top):
            cs  = CURRENCY_SYMBOLS.get(d.get('currency', 'UZS'), "so'm")
            ov  = " ⚠️" if (d.get('due_date') and d['due_date'] < today) else ""
            due = f"\n     📅 {d['due_date']}{ov}" if d.get('due_date') else ""
            txt += f"{medals[i]} <b>{d['name']}</b>\n     {int(d['remaining']):,} {cs}{due}\n\n".replace(',', ' ')
        await q.message.reply_text(txt, parse_mode='HTML')
        return

    if data == "monthly_chart":
        added_data, paid_data = db.get_monthly_chart_data(cid, 6)
        if not added_data:
            await q.answer("📭 Ma'lumot yo'q!", show_alert=True)
            return
        await q.answer()
        all_months = sorted(
            set(d['m'] for d in added_data) | set(d['m'] for d in paid_data)
        )
        added_map  = {d['m']: d['total'] for d in added_data}
        paid_map   = {d['m']: d['total'] for d in paid_data}
        added_vals = [added_map.get(m, 0) for m in all_months]
        paid_vals  = [paid_map.get(m, 0)  for m in all_months]
        mn         = ['Yan', 'Fev', 'Mar', 'Apr', 'May', 'Iyn',
                      'Iyl', 'Avg', 'Sen', 'Okt', 'Noy', 'Dek']
        labels     = []
        for m in all_months:
            p = m.split('-')
            labels.append(f"{mn[int(p[1])-1]}\n{p[0][2:]}" if len(p) == 2 else m)
        msg = await q.message.reply_text("⏳ Grafik tayyorlanmoqda...")
        try:
            path = generate_monthly_chart(added_vals, paid_vals, cid, labels)
            await msg.delete()
            if path and os.path.exists(path):
                with open(path, 'rb') as f:
                    await q.message.reply_photo(
                        photo=f,
                        caption="📅 <b>Oylik statistika grafigi</b>",
                        parse_mode='HTML'
                    )
                os.remove(path)
        except Exception as e:
            logger.error(f"Grafik xatosi: {e}")
            await msg.edit_text(f"❌ Grafik xatosi: {e}")
        return

    if data == "full_report":
        debts = db.get_active_debts(cid)
        if not debts:
            await q.answer("📭 Qarz yo'q!", show_alert=True)
            return
        await q.answer()
        m   = await q.message.reply_text("⏳ Tayyorlanmoqda...")
        try:
            res = generate_report(debts, cid)
            await m.delete()
            report_text = res['text']
            if len(report_text) > 1020:
                report_text = report_text[:1020] + "\n\n_...to'liq Excel/PDF da_"
            if res.get('chart_path') and os.path.exists(res['chart_path']):
                with open(res['chart_path'], 'rb') as f:
                    await q.message.reply_photo(photo=f, caption=report_text, parse_mode='Markdown')
                os.remove(res['chart_path'])
            else:
                await q.message.reply_text(report_text, parse_mode='Markdown')
        except Exception as e:
            logger.error(f"Full report xatosi: {e}")
            try:
                await m.delete()
            except Exception:
                pass
            await q.message.reply_text(f"❌ Xato: {e}")
        return

    if data == "get_excel":
        debts = db.get_active_debts(cid)
        if not debts:
            await q.answer("📭 Qarz yo'q!", show_alert=True)
            return
        await q.answer()
        m  = await q.message.reply_text("⏳ Excel tayyorlanmoqda...")
        try:
            ep = generate_excel(debts, db.get_all_debts(cid), cid)
            await m.delete()
            with open(ep, 'rb') as f:
                await q.message.reply_document(
                    document=f,
                    filename=f"qarz_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    caption=(
                        "📥 <b>Excel tayyor!</b>\n"
                        "<i>3 varaq: Faol qarzlar · Tarix · Statistika</i>"
                    ),
                    parse_mode='HTML'
                )
            os.remove(ep)
        except Exception as e:
            logger.error(f"Excel callback xatosi: {e}")
            try:
                await m.delete()
            except Exception:
                pass
            await q.message.reply_text(f"❌ Excel xatosi: {e}")
        return

    if data == "get_pdf":
        debts = db.get_active_debts(cid)
        if not debts:
            await q.answer("📭 Qarz yo'q!", show_alert=True)
            return
        await q.answer()
        m = await q.message.reply_text("⏳ PDF tayyorlanmoqda...")
        try:
            pp = generate_pdf(debts, cid)
            await m.delete()
            if pp and os.path.exists(pp):
                with open(pp, 'rb') as f:
                    await q.message.reply_document(
                        document=f,
                        filename=f"qarz_{datetime.now().strftime('%Y%m%d')}.pdf",
                        caption=(
                            "📄 <b>PDF hisobot tayyor!</b>\n"
                            "<i>Umumiy ko'rsatkichlar + Qarzdorlar jadvali</i>"
                        ),
                        parse_mode='HTML'
                    )
                os.remove(pp)
            else:
                await q.message.reply_text("❌ PDF yaratishda xato yuz berdi.")
        except Exception as e:
            logger.error(f"PDF xatosi: {e}")
            try:
                await m.delete()
            except Exception:
                pass
            await q.message.reply_text(f"❌ PDF xatosi: {e}")
        return

    if data == "how_debt":
        await q.answer()
        await q.message.edit_text(
            "📝 <b>Tez yozish usuli:</b>\n\n"
            "<code>Alisher 500000 qarz</code>\n"
            "<code>Nodira 1.5 mln karz</code>\n"
            "<code>Inomjon 6000</code>  ← kalit so'zsiz!\n"
            "<code>inomjon6000</code>   ← biriktirilgan!\n"
            "<code>Ali 150000қарз</code> ← kirilcha!\n\n"
            "✅ Bot o'zi o'qib, bazaga qo'shadi!",
            parse_mode='HTML'
        )
        return

    if data == "wizard_debt":
        await q.answer()
        await q.message.edit_text(
            "🪄 Bosqichma-bosqich kiritish uchun\n/qosh buyrug'ini yuboring.",
            parse_mode='HTML'
        )
        return

    if data.startswith("lang_"):
        await q.answer()
        lang = data.split("_")[1]
        db.set_setting(cid, 'lang', lang)
        names = {'uz': "O'zbek 🇺🇿", 'ru': "Русский 🇷🇺", 'en': "English 🇺🇸"}
        await q.message.edit_text(
            f"✅ Til o'zgartirildi: <b>{names.get(lang, lang)}</b>",
            parse_mode='HTML'
        )
        return


# ══════════════════════════════════════════════════════════════════════════════
#  BUYRUQLAR
# ══════════════════════════════════════════════════════════════════════════════
async def cmd_undo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid   = update.effective_chat.id
    entry = db.pop_undo(cid)
    if not entry:
        await update.message.reply_text(
            "❌ Bekor qilinadigan amal topilmadi.", reply_markup=MAIN_KB
        )
        return
    name = entry['data'].get('name', '?')
    desc = {
        'ADD_DEBT':    "Qarz qo'shish",
        'PAYMENT':     "To'lov",
        'DELETE_DEBT': "O'chirish",
        'UPDATE_DEBT': "Yangilash"
    }.get(entry['action_type'], entry['action_type'])
    await update.message.reply_text(
        f"↩️ <b>Bekor qilindi!</b>\n\nAmal: {desc}\nKim: <b>{name}</b>\nVaqt: {entry['created_at'][:16]}",
        parse_mode='HTML', reply_markup=MAIN_KB
    )

async def cmd_top(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid   = update.effective_chat.id
    top   = db.get_top_debtors(cid, 5)
    if not top:
        await update.message.reply_text("📭 Faol qarzlar yo'q!", reply_markup=MAIN_KB)
        return
    medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
    today  = datetime.now().strftime('%Y-%m-%d')
    txt    = "🏆 <b>TOP-5 QARZDORLAR</b>\n━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    for i, d in enumerate(top):
        cs  = CURRENCY_SYMBOLS.get(d.get('currency', 'UZS'), "so'm")
        ov  = " ⚠️ MUDDATI O'TDI" if (d.get('due_date') and d['due_date'] < today) else ""
        due = f"\n   📅 {d['due_date']}{ov}" if d.get('due_date') else ""
        txt += f"{medals[i]} <b>{d['name']}</b>\n   💸 {int(d['remaining']):,} {cs}{due}\n\n".replace(',', ' ')
    await update.message.reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)

async def cmd_overdue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid   = update.effective_chat.id
    debts = db.get_overdue_debts(cid)
    if not debts:
        await update.message.reply_text(
            "✅ Muddati o'tgan qarzlar yo'q! 🎉", reply_markup=MAIN_KB
        )
        return
    txt = "⚠️ <b>MUDDATI O'TGAN QARZLAR</b>\n━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    for d in debts:
        cs = CURRENCY_SYMBOLS.get(d.get('currency', 'UZS'), "so'm")
        txt += f"🔴 <b>{d['name']}</b>\n   💸 {int(d['remaining']):,} {cs}\n   📅 Muddat: {d['due_date']}\n\n".replace(',', ' ')
    await update.message.reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)

async def cmd_tarix(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    if not context.args:
        await update.message.reply_text(
            "Format: /tarix [ID]\nMisol: /tarix 5", reply_markup=MAIN_KB
        )
        return
    try:
        debt_id = int(context.args[0].replace('#', ''))
    except Exception:
        await update.message.reply_text("❌ ID noto'g'ri!", reply_markup=MAIN_KB)
        return
    debt, pays = db.get_payment_history(debt_id, cid)
    if not debt:
        await update.message.reply_text(f"❌ #{debt_id} topilmadi!", reply_markup=MAIN_KB)
        return
    cs  = CURRENCY_SYMBOLS.get(debt.get('currency', 'UZS'), "so'm")
    txt = (
        f"📋 <b>To'lov tarixi — {debt['name']}</b>\n━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"💸 Umumiy: <b>{int(debt['amount']):,} {cs}</b>\n"
        f"⏳ Qolgan: <b>{int(debt['remaining']):,} {cs}</b>\n\n"
    ).replace(',', ' ')
    if not pays:
        txt += "💳 Hali to'lovlar amalga oshirilmagan."
    else:
        for i, p in enumerate(pays, 1):
            txt += f"  {i}. {p['paid_at'][:10]} — <b>{int(p['amount']):,} {cs}</b>\n".replace(',', ' ')
        txt += f"\n✅ Jami: <b>{int(sum(p['amount'] for p in pays)):,} {cs}</b>".replace(',', ' ')
    await update.message.reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)

async def cmd_til(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid      = update.effective_chat.id
    cur_lang = db.get_setting(cid, 'lang', 'uz')
    await update.message.reply_text(
        "🌐 <b>Til tanlang / Выберите язык / Choose language:</b>",
        parse_mode='HTML', reply_markup=MAIN_KB
    )
    await update.message.reply_text("👇", reply_markup=InlineKeyboardMarkup([[
        InlineKeyboardButton("🇺🇿 O'zbek"   + (" ✅" if cur_lang == 'uz' else ""), callback_data="lang_uz"),
        InlineKeyboardButton("🇷🇺 Русский"  + (" ✅" if cur_lang == 'ru' else ""), callback_data="lang_ru"),
        InlineKeyboardButton("🇺🇸 English"  + (" ✅" if cur_lang == 'en' else ""), callback_data="lang_en"),
    ]]))

async def cmd_backup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    try:
        with open(db.db_path, 'rb') as f:
            data = f.read()
        bio      = io.BytesIO(data)
        bio.name = f"qarz_backup_{datetime.now().strftime('%Y%m%d_%H%M')}.db"
        debts    = db.get_active_debts(cid)
        await update.message.reply_document(
            document=bio, filename=bio.name,
            caption=(
                f"💾 <b>Zaxira nusxa tayyor!</b>\n"
                f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}\n"
                f"📊 {len(debts)} ta faol qarz\n\n"
                f"<i>Faylni xavfsiz joyda saqlang!</i>"
            ),
            parse_mode='HTML', reply_markup=MAIN_KB
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Backup xatosi: {e}", reply_markup=MAIN_KB)

async def cmd_eslatma(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    if len(context.args) < 2:
        await update.message.reply_text(
            "📨 Format: /eslatma [ID] [xabar]\n\nMisol:\n/eslatma 5 Salom, qarzingizni unutmang!",
            reply_markup=MAIN_KB
        )
        return
    try:
        debt_id = int(context.args[0].replace('#', ''))
    except Exception:
        await update.message.reply_text("❌ ID noto'g'ri!", reply_markup=MAIN_KB)
        return
    debt = db.get_debt_by_id(debt_id, cid)
    if not debt:
        await update.message.reply_text(f"❌ #{debt_id} topilmadi!", reply_markup=MAIN_KB)
        return
    msg_txt = ' '.join(context.args[1:])
    cs = CURRENCY_SYMBOLS.get(debt.get('currency', 'UZS'), "so'm")
    await update.message.reply_text(
        f"📨 <b>Eslatma: {debt['name']}</b>\n\n"
        f"💬 {msg_txt}\n\n"
        f"💸 Qolgan qarz: <b>{int(debt['remaining']):,} {cs}</b>".replace(',', ' '),
        parse_mode='HTML', reply_markup=MAIN_KB
    )

async def cmd_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cid = update.effective_chat.id
    if not update.message.document:
        await update.message.reply_text(
            "📥 <b>Excel/CSV import</b>\n\n"
            "Excel faylida ustunlar:\n"
            "<b>A:</b> Ism  |  <b>B:</b> Summa  |  <b>C:</b> Izoh\n"
            "<b>D:</b> Valyuta  |  <b>E:</b> Kategoriya\n\n"
            "<i>Faylni shu chatga yuboring (.xlsx yoki .csv)</i>",
            parse_mode='HTML', reply_markup=MAIN_KB
        )
        return

    doc   = update.message.document
    fname = doc.file_name or ''
    if not (fname.endswith('.xlsx') or fname.endswith('.csv')):
        await update.message.reply_text("❌ Faqat .xlsx yoki .csv fayl!", reply_markup=MAIN_KB)
        return

    m = await update.message.reply_text("⏳ Import qilinmoqda...")
    try:
        file = await doc.get_file()
        tmp  = os.path.join(tempfile.gettempdir(), f"import_{cid}_{os.path.basename(fname)}")
        await file.download_to_drive(tmp)
        rows = []
        if fname.endswith('.xlsx'):
            from openpyxl import load_workbook
            wb2 = load_workbook(tmp, data_only=True)
            ws  = wb2.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    rows.append({
                        'name':     row[0],
                        'amount':   row[1] or 0,
                        'note':     row[2] or '',
                        'currency': row[3] or 'UZS',
                        'category': row[4] or ''
                    })
        else:
            import csv
            with open(tmp, encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    rows.append(row)

        cnt = db.bulk_import(cid, rows, added_by="@import")
        os.remove(tmp)
        await m.delete()
        await update.message.reply_text(
            f"✅ <b>Import muvaffaqiyatli!</b>\n\n"
            f"📊 {cnt} ta qarz qo'shildi\n"
            f"📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            parse_mode='HTML', reply_markup=MAIN_KB
        )
    except Exception as e:
        try:
            await m.delete()
        except Exception:
            pass
        await update.message.reply_text(f"❌ Import xatosi: {e}", reply_markup=MAIN_KB)

async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    txt = (
        "❓ <b>YORDAM — v5.2</b>\n"
        "━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        "📥 <b>Qarz yozish:</b>\n"
        "  <code>Alisher 500000 qarz</code>\n"
        "  <code>Nodira 1.5 mln karz</code>\n"
        "  <code>@sardor 300000 қарз</code> ← kirilcha\n"
        "  <code>inomjon6000</code>   ← biriktirilgan\n"
        "  <code>Ali 100000қарз</code> ← kirilcha biriktirilgan\n"
        "  <code>Inomjon 6000</code>   ← kalit so'zsiz\n\n"
        "💳 <b>To'lov yozish:</b>\n"
        "  <code>Alisher 200000 berdi</code>\n"
        "  <code>Nodira 500000 to'ladi</code>\n"
        "  <code>Bobur qaytardi 100000</code>\n\n"
        "📌 <b>Buyruqlar:</b>\n"
        "  /start    — Boshlash\n"
        "  /qosh     — Qarz qo'shish (batafsil)\n"
        "  /qarzlar  — Ro'yxat\n"
        "  /top      — TOP-5 🏆\n"
        "  /tarix 5  — #5 to'lov tarixi\n"
        "  /muddat   — Muddati o'tganlar ⚠️\n"
        "  /eslatma 5 [xabar]\n"
        "  /bekor    — Undo ↩️\n"
        "  /backup   — Zaxira 💾\n"
        "  /import   — Excel/CSV import\n"
        "  /til      — Til tanlash 🌐\n"
        "  /panel    — Tezkor panel 📱\n"
        "  /audit    — Oxirgi amallar jurnali\n"
        "  /eslatmakun 7 — Eslatma har 7 kun\n"
        "  /bonus 5  — Muddatdan oldin +5% hisob\n"
        "  /kurs     — .env kurslari\n\n"
        "🔔 <b>Eslatmalar:</b> sozlangan interval (default 10 kun), tekshiruv ~1 soat"
    )
    await update.message.reply_text(txt, parse_mode='HTML', reply_markup=MAIN_KB)


async def cmd_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("⌨️ Klaviatura yangilanmoqda...", reply_markup=ReplyKeyboardRemove())
    await update.message.reply_text(
        "📱 Menyu qayta chiqarildi. Pastdagi tugmalardan foydalaning.",
        reply_markup=MAIN_KB
    )


# ══════════════════════════════════════════════════════════════════════════════
#  OVOZLI XABAR
# ══════════════════════════════════════════════════════════════════════════════
async def handle_voice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        import speech_recognition as sr
        from pydub import AudioSegment
    except ImportError:
        await update.message.reply_text(
            "❌ Ovozli xabar uchun:\n<code>pip install SpeechRecognition pydub</code>",
            parse_mode='HTML', reply_markup=MAIN_KB
        )
        return

    recognizer = sr.Recognizer()
    voice      = await update.message.voice.get_file()
    _td = tempfile.gettempdir()
    vp = os.path.join(_td, f"{voice.file_id}.ogg")
    wp = os.path.join(_td, f"{voice.file_id}.wav")
    pm = await update.message.reply_text("🎤 <b>Ovoz tahlil qilinmoqda...</b>", parse_mode='HTML')
    await voice.download_to_drive(vp)
    try:
        AudioSegment.from_ogg(vp).export(wp, format="wav")
        with sr.AudioFile(wp) as s:
            text = recognizer.recognize_google(recognizer.record(s), language="uz-UZ")
        await pm.delete()
        await update.message.reply_text(
            f"🎤 <i>Eshitildi:</i> <b>{text}</b>", parse_mode='HTML'
        )
        await do_process(update, text, update.effective_chat.id)
    except sr.UnknownValueError:
        await pm.edit_text("❌ Ovozni tushunmadim. Aniqroq gapiring.")
    except Exception as e:
        logger.error(f"Voice error: {e}")
        await pm.edit_text("❌ Xato yuz berdi. FFmpeg o'rnatilganini tekshiring.")
    finally:
        for p in [vp, wp]:
            if os.path.exists(p):
                os.remove(p)


# ══════════════════════════════════════════════════════════════════════════════
#  ESLATMA (interval chat_settings: reminder_days, reminder_last_sent)
# ══════════════════════════════════════════════════════════════════════════════
async def auto_reminder(context: ContextTypes.DEFAULT_TYPE):
    today = datetime.now().strftime('%Y-%m-%d')
    for cid in db.get_all_chat_ids():
        try:
            interval = int(db.get_setting(cid, 'reminder_days', '10') or 10)
        except ValueError:
            interval = 10
        last = (db.get_setting(cid, 'reminder_last_sent', '') or '').strip()
        if last:
            try:
                ld = datetime.strptime(last[:10], '%Y-%m-%d').date()
                delta = (datetime.now().date() - ld).days
                if delta < interval:
                    continue
            except ValueError:
                pass

        debts   = db.get_active_debts(cid)
        overdue = db.get_overdue_debts(cid)
        if not debts:
            continue

        lines = "\n".join(
            f"  • <b>{d['name']}</b>: {int(d['remaining']):,} "
            f"{CURRENCY_SYMBOLS.get(d.get('currency', 'UZS'), '')}".replace(',', ' ')
            for d in debts[:8]
        )
        sums_line = _list_remaining_header(debts).strip()
        title = T(cid, 'reminder_title', db, days=interval)
        txt = (
            f"{title} 📢\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"{T(cid, 'reminder_active', db, n=len(debts))}\n{lines}\n\n"
            f"{sums_line}"
        )
        if overdue:
            txt += T(cid, 'reminder_overdue', db, m=len(overdue))
        try:
            await context.bot.send_message(cid, txt, parse_mode='HTML')
            db.set_setting(cid, 'reminder_last_sent', today)
        except Exception as e:
            logger.warning(f"Eslatma xatosi ({cid}): {e}")


async def job_auto_backup(context: ContextTypes.DEFAULT_TYPE):
    if not ADMIN_IDS:
        return
    try:
        with open(db.db_path, 'rb') as f:
            raw = f.read()
    except OSError as e:
        logger.warning("Zaxira o'qilmadi: %s", e)
        return
    cap = f"💾 <b>Avtomatik zaxira</b>\n📅 {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    for aid in ADMIN_IDS:
        try:
            bio = io.BytesIO(raw)
            bio.name = f"qarz_auto_{datetime.now().strftime('%Y%m%d_%H%M')}.db"
            await context.bot.send_document(aid, document=bio, caption=cap, parse_mode='HTML')
        except Exception as e:
            logger.warning("Admin %s ga zaxira yuborilmadi: %s", aid, e)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.error("Handlerda xato", exc_info=context.error)
    if not isinstance(update, Update):
        return
    try:
        if update.effective_message:
            await update.effective_message.reply_text(
                T(update.effective_chat.id, 'err_internal', db) if update.effective_chat
                else "❌ Xato",
                reply_markup=MAIN_KB,
            )
        elif update.callback_query and update.callback_query.message:
            await update.callback_query.message.reply_text(
                T(update.callback_query.message.chat_id, 'err_internal', db),
                reply_markup=MAIN_KB,
            )
    except Exception as ex:
        logger.warning("error_handler: %s", ex)


async def post_init(application: Application) -> None:
    if _disable_webhook:
        # Polling mode requires webhook to be disabled on Telegram side.
        await application.bot.delete_webhook(drop_pending_updates=True)
    cmds = [
        BotCommand("start", "Boshlash"),
        BotCommand("panel", "Tezkor panel"),
        BotCommand("qarzlar", "Ro'yxat"),
        BotCommand("yordam", "Yordam"),
        BotCommand("audit", "Audit jurnali"),
        BotCommand("eslatmakun", "Eslatma intervali"),
        BotCommand("bonus", "Muddatdan oldin bonus %"),
        BotCommand("kurs", "Valyuta kurslari"),
        BotCommand("backup", "Zaxira nusxa"),
        BotCommand("menu", "Menyuni qayta ko'rsatish"),
    ]
    await application.bot.set_my_commands(cmds)
    if MINI_APP_URL:
        await application.bot.set_chat_menu_button(
            menu_button=MenuButtonWebApp(
                text=T(0, 'miniapp_btn', db),
                web_app=WebAppInfo(url=MINI_APP_URL),
            )
        )
    else:
        await application.bot.set_chat_menu_button(menu_button=MenuButtonCommands())


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    _start_health_server_if_needed()
    app = Application.builder().token(BOT_TOKEN).post_init(post_init).build()

    wizard = ConversationHandler(
        entry_points=[
            CommandHandler("qosh", wizard_start),
            CommandHandler("add",  wizard_start),
        ],
        states={
            S_NAME:     [MessageHandler(filters.TEXT & ~filters.COMMAND, w_name)],
            S_CURRENCY: [MessageHandler(filters.TEXT & ~filters.COMMAND, w_currency)],
            S_AMOUNT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, w_amount)],
            S_CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, w_category)],
            S_DUE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, w_due)],
            S_NOTE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, w_note)],
        },
        fallbacks=[CommandHandler("bekor", wizard_cancel)],
        allow_reentry=True,
        per_message=False,
    )

    app.add_handler(wizard)
    app.add_handler(CommandHandler("start",   cmd_start))
    app.add_handler(CommandHandler("qarzlar", lambda u, c: send_list(u, 0)))
    app.add_handler(CommandHandler("list",    lambda u, c: send_list(u, 0)))
    app.add_handler(CommandHandler("top",     cmd_top))
    app.add_handler(CommandHandler("tarix",   cmd_tarix))
    app.add_handler(CommandHandler("muddat",  cmd_overdue))
    app.add_handler(CommandHandler("bekor",   cmd_undo))
    app.add_handler(CommandHandler("eslatma", cmd_eslatma))
    app.add_handler(CommandHandler("til",     cmd_til))
    app.add_handler(CommandHandler("backup",  cmd_backup))
    app.add_handler(CommandHandler("import",  cmd_import))
    app.add_handler(CommandHandler("yordam",  cmd_help))
    app.add_handler(CommandHandler("help",    cmd_help))
    app.add_handler(CommandHandler("menu",    cmd_menu))
    app.add_handler(CommandHandler("panel",   cmd_panel))
    app.add_handler(CommandHandler("audit",   cmd_audit))
    app.add_handler(CommandHandler("eslatmakun", cmd_eslatmakun))
    app.add_handler(CommandHandler("bonus",   cmd_bonus))
    app.add_handler(CommandHandler("kurs",    cmd_kurs))

    app.add_error_handler(error_handler)
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_handler(MessageHandler(filters.Document.ALL, cmd_import))
    app.add_handler(MessageHandler(filters.VOICE, handle_voice))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    app.job_queue.run_repeating(auto_reminder, interval=3600, first=90)
    if ADMIN_IDS:
        app.job_queue.run_daily(job_auto_backup, time=timeofday(hour=3, minute=7))

    print("\n" + "═" * 56)
    print("  🏦  QARZ HISOBCHI BOT v5.2")
    print("═" * 56)
    print("  ✅  Audit · panel · eslatma intervali · bonus · kurs")
    print("  ✅  Webhook: WEBHOOK_URL o'rnatilsa")
    print("  ✅  Mini App: MINI_APP_URL")
    print("  ⛔  To'xtatish: Ctrl+C")
    print("═" * 56 + "\n")

    if WEBHOOK_URL:
        wh = WEBHOOK_URL.rstrip("/") + "/" + WEBHOOK_PATH
        print("  🌐 Webhook:", wh)
        app.run_webhook(
            listen="0.0.0.0",
            port=WEBHOOK_PORT,
            url_path=WEBHOOK_PATH,
            webhook_url=wh,
            allowed_updates=Update.ALL_TYPES,
        )
    else:
        app.run_polling(
            allowed_updates=Update.ALL_TYPES,
            drop_pending_updates=True,
        )


if __name__ == '__main__':
    main()
