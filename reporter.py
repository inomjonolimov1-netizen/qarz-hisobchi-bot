"""
Hisobot generatori v5.1 — TUZATILGAN
- Excel: ma'lumotlar ko'rinadigan (to'q rangda qora matn)
- PDF: to'g'irlangan, Unicode xatoliklaridan himoya
- Grafik: chiroyli, aniq, zamonaviy dizayn
"""
import os, tempfile
from datetime import datetime


# ─── Yordamchi ─────────────────────────────────────────────────────────────
def fmt(n):
    """Sonni bo'sh joy bilan formatlash: 1 500 000"""
    return f"{int(n):,}".replace(',', ' ')

def fmts(n):
    """Qisqa format: 1.5M, 500K"""
    if n >= 1_000_000: return f"{n/1_000_000:.1f}M"
    if n >= 1_000:     return f"{n/1000:.0f}K"
    return str(int(n))

def pbar(pct, n=10):
    f = int(n * pct / 100)
    return "🟩" * f + "⬜" * (n - f)

CSYM = {'UZS': "so'm", 'USD': '$', 'RUB': '₽'}


# ─── Matnli hisobot ────────────────────────────────────────────────────────
def generate_report(debts, chat_id, prefix=''):
    now   = datetime.now()
    t_amt = sum(d['amount']    for d in debts)
    t_rem = sum(d['remaining'] for d in debts)
    t_pay = t_amt - t_rem

    text  = prefix or f"📊 *Qarz Hisoboti — {now.strftime('%d.%m.%Y %H:%M')}*\n\n"
    text += f"👥 Jami qarzdorlar: *{len(debts)} kishi*\n"
    text += f"💰 Umumiy qarz: *{fmt(t_amt)} so'm*\n"
    if t_pay > 0:
        text += f"✅ To'langan: *{fmt(t_pay)} so'm*\n"
        text += f"⏳ Qolgan: *{fmt(t_rem)} so'm*\n"
    text += "\n━━━━━━━━━━━━━━━━━━━━\n\n"

    for i, d in enumerate(sorted(debts, key=lambda x: x['remaining'], reverse=True), 1):
        paid = d['amount'] - d['remaining']
        pct  = int(paid / d['amount'] * 100) if d['amount'] > 0 else 0
        cur  = d.get('currency', 'UZS')
        csym = CSYM.get(cur, cur)
        dt   = d.get('created_at', '')[:10]
        cat  = f" [{d['category']}]" if d.get('category') else ''
        due  = f" · Muddat: {d['due_date']}" if d.get('due_date') else ''
        text += f"*{i}. {d['name']}*{cat}\n"
        text += f"   {fmt(d['amount'])} {csym}\n"
        if paid > 0:
            text += f"   Tolandi: {fmt(paid)} {csym} ({pct}%)\n"
            text += f"   Qoldi: {fmt(d['remaining'])} {csym}\n"
        text += f"   {dt}{due} | #{d['id']}\n\n"

    text += f"━━━━━━━━━━━━━━━━━━━━\n_{now.strftime('%d.%m.%Y %H:%M')}_"

    chart = None
    try:
        chart = generate_main_chart(
            sorted(debts, key=lambda x: x['remaining'], reverse=True), chat_id
        )
    except Exception as e:
        print(f"Chart xatosi: {e}")

    return {'text': text, 'chart_path': chart,
            'total_debt': t_amt, 'total_remaining': t_rem}


# ─── Asosiy diagramma (YANGILANGAN — chiroyli va aniq) ─────────────────────
def generate_main_chart(debts, chat_id):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    import matplotlib.ticker as ticker
    import numpy as np

    if not debts:
        return None

    # ── Ranglar ──────────────────────────────────────────────────────────────
    BG      = '#0d1117'   # fon
    CARD    = '#161b22'   # karta foni
    RED     = '#ff4d6d'   # umumiy qarz
    GREEN   = '#00d084'   # to'langan
    BLUE    = '#4f8ef7'   # qolgan
    YELLOW  = '#ffd60a'   # sariq
    GRAY    = '#8b949e'   # kulrang
    WHITE   = '#e6edf3'   # oq matn

    names      = [d['name'][:16] for d in debts]
    amounts    = [float(d['amount'])    for d in debts]
    remainings = [float(d['remaining']) for d in debts]
    paid_list  = [float(d['amount'] - d['remaining']) for d in debts]
    n = len(names)

    fig = plt.figure(figsize=(16, 12), facecolor=BG)
    fig.suptitle(
        f"QARZ HISOBOTI  ·  {datetime.now().strftime('%d.%m.%Y')}",
        color=WHITE, fontsize=18, fontweight='bold', y=0.98
    )

    gs = fig.add_gridspec(2, 2, hspace=0.38, wspace=0.32,
                          left=0.08, right=0.95, top=0.93, bottom=0.07)

    # ── 1. Gorizontal bar chart ───────────────────────────────────────────────
    ax1 = fig.add_subplot(gs[0, :])
    ax1.set_facecolor(CARD)
    y  = np.arange(n)
    bh = 0.28

    bars1 = ax1.barh(y + bh,     amounts,    bh, color=RED,   alpha=0.90, label="Umumiy qarz",  zorder=3)
    bars2 = ax1.barh(y,           remainings, bh, color=BLUE,  alpha=0.90, label="Qolgan",       zorder=3)
    if any(p > 0 for p in paid_list):
        bars3 = ax1.barh(y - bh, paid_list,  bh, color=GREEN, alpha=0.90, label="To'langan",    zorder=3)

    ax1.set_yticks(y)
    ax1.set_yticklabels(names, color=WHITE, fontsize=11, fontweight='bold')
    ax1.set_title("Qarzdorlik holati", color=WHITE, fontsize=14, fontweight='bold', pad=12)
    ax1.tick_params(colors=GRAY, labelsize=9)
    ax1.xaxis.set_major_formatter(
        ticker.FuncFormatter(lambda x, _: f"{x/1e6:.1f}M" if x >= 1e6 else f"{x/1e3:.0f}K" if x >= 1e3 else str(int(x)))
    )
    for sp in ax1.spines.values():
        sp.set_color('#30363d')
    ax1.grid(axis='x', color='#30363d', alpha=0.6, zorder=0)
    ax1.legend(facecolor=CARD, edgecolor='#30363d', labelcolor=WHITE,
               fontsize=10, loc='lower right', framealpha=0.9)

    # Qiymatlarni yozish
    for bar in bars1:
        w = bar.get_width()
        if w > 0:
            ax1.text(w * 1.01, bar.get_y() + bar.get_height()/2,
                     fmts(w), va='center', ha='left', color=RED, fontsize=8, fontweight='bold')
    for bar in bars2:
        w = bar.get_width()
        if w > 0:
            ax1.text(w * 1.01, bar.get_y() + bar.get_height()/2,
                     fmts(w), va='center', ha='left', color=BLUE, fontsize=8, fontweight='bold')

    # ── 2. Pie chart ─────────────────────────────────────────────────────────
    ax2 = fig.add_subplot(gs[1, 0])
    ax2.set_facecolor(CARD)

    pie_colors = [RED, BLUE, GREEN, YELLOW, '#a78bfa', '#fb923c', '#34d399', '#60a5fa', '#f472b6', '#facc15']
    wa = remainings if any(r > 0 for r in remainings) else amounts
    lp = [f"{names[i][:10]}\n{fmts(wa[i])}" for i in range(n)]

    wedges, texts, autotexts = ax2.pie(
        wa, labels=lp, colors=pie_colors[:n],
        autopct='%1.0f%%', startangle=140, pctdistance=0.72,
        wedgeprops=dict(width=0.65, edgecolor=BG, linewidth=2),
        textprops={'color': WHITE, 'fontsize': 8}
    )
    for at in autotexts:
        at.set_fontsize(9)
        at.set_fontweight('bold')
        at.set_color(WHITE)
    ax2.set_title("Qarz ulushi (qolgan)", color=WHITE, fontsize=13, fontweight='bold', pad=12)

    # ── 3. Ko'rsatkichlar paneli ─────────────────────────────────────────────
    ax3 = fig.add_subplot(gs[1, 1])
    ax3.set_facecolor(CARD)
    ax3.axis('off')
    ax3.set_title("Umumiy ko'rsatkichlar", color=WHITE, fontsize=13, fontweight='bold', pad=12)

    t_o = sum(amounts)
    t_r = sum(remainings)
    t_p = sum(paid_list)
    pp  = t_p / t_o * 100 if t_o > 0 else 0

    # Progress bar
    pct_fill = int(pp)
    bar_str  = "█" * int(pct_fill / 5) + "░" * (20 - int(pct_fill / 5))

    stats = [
        ("👥  Qarzdorlar",  f"{n} kishi",       WHITE),
        ("💸  Umumiy qarz", fmts(t_o),           RED),
        ("✅  To'langan",   fmts(t_p),           GREEN),
        ("⏳  Qolgan",      fmts(t_r),           BLUE),
        ("📈  To'lov %",    f"{pp:.1f}%",        YELLOW),
    ]

    yp = 0.88
    for label, val, color in stats:
        ax3.text(0.05, yp, label, transform=ax3.transAxes,
                 color=GRAY, fontsize=11, va='top')
        ax3.text(0.95, yp, val,  transform=ax3.transAxes,
                 color=color, fontsize=12, fontweight='bold', va='top', ha='right')
        ax3.plot([0.05, 0.95], [yp - 0.035, yp - 0.035],
                 color='#30363d', lw=0.8, transform=ax3.transAxes, clip_on=False)
        yp -= 0.17

    # Progress bar ko'rsatish
    yp -= 0.02
    ax3.text(0.05, yp, f"Progress: {bar_str} {pp:.0f}%",
             transform=ax3.transAxes, color=GREEN, fontsize=9, va='top',
             fontfamily='monospace')

    tmp = tempfile.NamedTemporaryFile(
        suffix=f'_main_{chat_id}.png', delete=False, dir=tempfile.gettempdir()
    )
    plt.savefig(tmp.name, dpi=130, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close()
    return tmp.name


# ─── Oylik grafik (YANGILANGAN) ────────────────────────────────────────────
def generate_monthly_chart(monthly_added, monthly_paid, chat_id, months_labels):
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    BG    = '#0d1117'
    CARD  = '#161b22'
    RED   = '#ff4d6d'
    GREEN = '#00d084'
    WHITE = '#e6edf3'
    GRAY  = '#8b949e'

    fig, axes = plt.subplots(1, 3, figsize=(18, 7), facecolor=BG)
    fig.suptitle(
        f"OYLIK STATISTIKA  ·  {datetime.now().strftime('%Y')}",
        color=WHITE, fontsize=16, fontweight='bold', y=0.98
    )

    added_vals = [v / 1_000_000 for v in monthly_added]
    paid_vals  = [v / 1_000_000 for v in monthly_paid]
    x = np.arange(len(months_labels))
    w = 0.38

    # ── 1. Grouped bar chart ──────────────────────────────────────────────────
    ax1 = axes[0]
    ax1.set_facecolor(CARD)
    b1 = ax1.bar(x - w/2, added_vals, w, color=RED,   alpha=0.9, label="Qo'shilgan",  zorder=3)
    b2 = ax1.bar(x + w/2, paid_vals,  w, color=GREEN, alpha=0.9, label="To'langan",   zorder=3)
    ax1.set_xticks(x)
    ax1.set_xticklabels(months_labels, color=WHITE, fontsize=9)
    ax1.set_ylabel("mln so'm", color=GRAY, fontsize=10)
    ax1.set_title("Oylik qarz va to'lovlar", color=WHITE, fontsize=12, fontweight='bold')
    ax1.legend(facecolor=CARD, edgecolor='#30363d', labelcolor=WHITE, fontsize=9)
    ax1.tick_params(colors=GRAY)
    for sp in ax1.spines.values(): sp.set_color('#30363d')
    ax1.grid(axis='y', color='#30363d', alpha=0.5, zorder=0)
    ax1.yaxis.label.set_color(GRAY)

    # ── 2. Trend line ─────────────────────────────────────────────────────────
    ax2 = axes[1]
    ax2.set_facecolor(CARD)
    ax2.plot(months_labels, added_vals, 'o-', color=RED,   linewidth=2.5, markersize=8,
             label="Qo'shilgan", zorder=3, markerfacecolor='white', markeredgecolor=RED, markeredgewidth=2)
    ax2.plot(months_labels, paid_vals,  's-', color=GREEN, linewidth=2.5, markersize=8,
             label="To'langan",  zorder=3, markerfacecolor='white', markeredgecolor=GREEN, markeredgewidth=2)
    ax2.fill_between(range(len(months_labels)), added_vals, alpha=0.12, color=RED)
    ax2.fill_between(range(len(months_labels)), paid_vals,  alpha=0.12, color=GREEN)
    ax2.set_xticks(range(len(months_labels)))
    ax2.set_xticklabels(months_labels, color=WHITE, fontsize=9)
    ax2.set_ylabel("mln so'm", color=GRAY, fontsize=10)
    ax2.set_title("Trend chizig'i", color=WHITE, fontsize=12, fontweight='bold')
    ax2.legend(facecolor=CARD, edgecolor='#30363d', labelcolor=WHITE, fontsize=9)
    ax2.tick_params(colors=GRAY)
    for sp in ax2.spines.values(): sp.set_color('#30363d')
    ax2.grid(color='#30363d', alpha=0.5, zorder=0)

    # Qiymatlarni yozish
    for i, (av, pv) in enumerate(zip(added_vals, paid_vals)):
        if av > 0:
            ax2.annotate(f"{av:.1f}M", (i, av), textcoords="offset points",
                         xytext=(0, 8), ha='center', color=RED, fontsize=7, fontweight='bold')
        if pv > 0:
            ax2.annotate(f"{pv:.1f}M", (i, pv), textcoords="offset points",
                         xytext=(0, -14), ha='center', color=GREEN, fontsize=7, fontweight='bold')

    # ── 3. Farq (Net) chart ───────────────────────────────────────────────────
    ax3 = axes[2]
    ax3.set_facecolor(CARD)
    net = [a - p for a, p in zip(added_vals, paid_vals)]
    bar_colors = [RED if v > 0 else GREEN for v in net]
    ax3.bar(x, net, color=bar_colors, alpha=0.85, zorder=3)
    ax3.axhline(0, color=GRAY, linewidth=1.2, linestyle='--', alpha=0.7)
    ax3.set_xticks(x)
    ax3.set_xticklabels(months_labels, color=WHITE, fontsize=9)
    ax3.set_ylabel("mln so'm (farq)", color=GRAY, fontsize=10)
    ax3.set_title("Sof o'zgarish", color=WHITE, fontsize=12, fontweight='bold')
    ax3.tick_params(colors=GRAY)
    for sp in ax3.spines.values(): sp.set_color('#30363d')
    ax3.grid(axis='y', color='#30363d', alpha=0.5, zorder=0)

    for i, v in enumerate(net):
        ax3.text(i, v + (0.02 if v >= 0 else -0.04),
                 f"{v:+.1f}M", ha='center', va='bottom' if v >= 0 else 'top',
                 color=WHITE, fontsize=8, fontweight='bold')

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    tmp = tempfile.NamedTemporaryFile(
        suffix=f'_monthly_{chat_id}.png', delete=False, dir=tempfile.gettempdir()
    )
    plt.savefig(tmp.name, dpi=130, bbox_inches='tight',
                facecolor=BG, edgecolor='none')
    plt.close()
    return tmp.name


# ─── Excel (TUZATILGAN — ma'lumotlar ko'rinadigan) ─────────────────────────
def generate_excel(active_debts, all_debts, chat_id):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Ranglar (hex, # belgisisiz) ───────────────────────────────────────────
    C_RED     = "E94560"   # sarlavha qizil
    C_DARK    = "1A1A2E"   # to'q ko'k
    C_MID     = "16213E"   # o'rta ko'k
    C_LIGHT   = "0F3460"   # ochroq ko'k
    C_GREEN   = "0A3D2E"   # yashil (to'langan)
    C_WHITE   = "FFFFFF"   # oq
    C_YELLOW  = "FFD700"   # sariq (jami)
    C_LGRAY   = "F5F5F5"   # açiq kulrang (data rows)
    C_MGRAY   = "E8E8E8"   # alt row

    def thin_border(color="CCCCCC"):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def header_style(ws, row, col_count, bg=C_RED, fg=C_WHITE, size=11):
        for c in range(1, col_count + 1):
            cell = ws.cell(row, c)
            cell.font      = Font(bold=True, color=fg, size=size)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = thin_border("888888")

    def data_style(ws, row, col_count, alt=False):
        """Ma'lumot qatorlari — qora matn, oq/kulrang fon"""
        bg = C_MGRAY if alt else C_LGRAY
        for c in range(1, col_count + 1):
            cell = ws.cell(row, c)
            cell.font      = Font(color="111111", size=10)   # QORA matn
            cell.fill      = PatternFill("solid", fgColor=bg) # OCH fon
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = thin_border()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1: Faol qarzlar
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Faol Qarzlar"
    ws1.sheet_view.showGridLines = False

    # Sarlavha
    ws1.merge_cells("A1:I1")
    c = ws1["A1"]
    c.value     = f"FAOL QARZDORLAR — {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    c.font      = Font(bold=True, color=C_WHITE, size=15)
    c.fill      = PatternFill("solid", fgColor=C_RED)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 38

    # Ustun boshliqlari
    heads = ["№", "Ism", "Qarz (so'm)", "To'langan", "Qolgan", "To'lov %", "Valyuta", "Kategoriya", "Muddat"]
    ws1.row_dimensions[2].height = 28
    for ci, h in enumerate(heads, 1):
        ws1.cell(2, ci).value = h
    header_style(ws1, 2, len(heads), bg=C_DARK)

    # Ma'lumotlar
    for i, d in enumerate(active_debts, 1):
        row  = i + 2
        paid = d['amount'] - d['remaining']
        pct  = round(paid / d['amount'] * 100, 1) if d['amount'] > 0 else 0
        data = [
            i,
            d['name'],
            d['amount'],
            paid,
            d['remaining'],
            pct,
            d.get('currency', 'UZS'),
            d.get('category', '') or '-',
            d.get('due_date', '') or '-'
        ]
        for ci, v in enumerate(data, 1):
            ws1.cell(row, ci).value = v
        data_style(ws1, row, len(heads), i % 2 == 0)

        # Raqam formati
        for ci in [3, 4, 5]:
            ws1.cell(row, ci).number_format = '#,##0'
        ws1.cell(row, 6).number_format = '0.0'
        ws1.row_dimensions[row].height = 22

        # Agar muddati o'tgan bo'lsa — sariq rang
        today = datetime.now().strftime('%Y-%m-%d')
        due = d.get('due_date', '')
        if due and due < today and d['remaining'] > 0:
            for ci in range(1, len(heads) + 1):
                ws1.cell(row, ci).fill = PatternFill("solid", fgColor="FFF3CD")
                ws1.cell(row, ci).font = Font(color="7B3F00", size=10)

    # Jami qator
    tr = len(active_debts) + 3
    total_data = {
        1: "JAMI",
        3: sum(d['amount'] for d in active_debts),
        4: sum(d['amount'] - d['remaining'] for d in active_debts),
        5: sum(d['remaining'] for d in active_debts),
    }
    for ci in range(1, len(heads) + 1):
        cell = ws1.cell(tr, ci)
        cell.value  = total_data.get(ci, '')
        cell.font   = Font(bold=True, color="111111", size=12)
        cell.fill   = PatternFill("solid", fgColor=C_YELLOW)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border("888888")
    for ci in [3, 4, 5]:
        ws1.cell(tr, ci).number_format = '#,##0'
    ws1.row_dimensions[tr].height = 28

    # Ustun kengliklari
    for ci, w in enumerate([4, 20, 16, 16, 16, 10, 10, 16, 14], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2: To'liq tarix
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Tarix")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:H1")
    t2 = ws2["A1"]
    t2.value     = f"BARCHA QARZLAR TARIXI — {datetime.now().strftime('%d.%m.%Y')}"
    t2.font      = Font(bold=True, color=C_WHITE, size=14)
    t2.fill      = PatternFill("solid", fgColor=C_LIGHT)
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 35

    heads2 = ["№", "Ism", "Qarz", "To'langan", "Qolgan", "Holat", "Sana", "Kim qo'shdi"]
    ws2.row_dimensions[2].height = 25
    for ci, h in enumerate(heads2, 1):
        ws2.cell(2, ci).value = h
    header_style(ws2, 2, len(heads2), bg=C_LIGHT)

    for i, d in enumerate(all_debts, 1):
        row  = i + 2
        data = [
            i, d['name'], d['amount'], d.get('paid', 0),
            d['remaining'], d.get('status', ''),
            d.get('created_at', '')[:10], d.get('added_by', '')
        ]
        for ci, v in enumerate(data, 1):
            ws2.cell(row, ci).value = v
        data_style(ws2, row, len(heads2), i % 2 == 0)

        if d.get('status') == "To'langan":
            for ci in range(1, len(heads2) + 1):
                ws2.cell(row, ci).fill = PatternFill("solid", fgColor="D4EDDA")
                ws2.cell(row, ci).font = Font(color="155724", size=10)

        for ci in [3, 4, 5]:
            ws2.cell(row, ci).number_format = '#,##0'
        ws2.row_dimensions[row].height = 20

    for ci, w in enumerate([4, 20, 16, 16, 16, 14, 12, 18], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3: Statistika + Grafik
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Statistika")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:E1")
    t3 = ws3["A1"]
    t3.value     = "STATISTIKA VA GRAFIKLAR"
    t3.font      = Font(bold=True, color=C_WHITE, size=14)
    t3.fill      = PatternFill("solid", fgColor=C_RED)
    t3.alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 35

    # Statistika jadvali
    header_style(ws3, 2, 2, bg=C_DARK)
    ws3.cell(2, 1).value = "Ko'rsatkich"
    ws3.cell(2, 2).value = "Qiymat"

    t_amt = sum(d['amount']    for d in active_debts)
    t_rem = sum(d['remaining'] for d in active_debts)
    t_pay = t_amt - t_rem

    stats_data = [
        ("Jami qarzdorlar",     f"{len(active_debts)} kishi"),
        ("Umumiy qarz (so'm)",  t_amt),
        ("To'langan (so'm)",    t_pay),
        ("Qolgan (so'm)",       t_rem),
        ("To'lov foizi",        f"{int(t_pay/t_amt*100) if t_amt > 0 else 0}%"),
    ]
    for i, (lb, v) in enumerate(stats_data, 3):
        ws3.cell(i, 1).value = lb
        ws3.cell(i, 2).value = v
        alt = i % 2 == 0
        for ci in [1, 2]:
            cell = ws3.cell(i, ci)
            cell.fill      = PatternFill("solid", fgColor=C_MGRAY if alt else C_LGRAY)
            cell.font      = Font(color="111111", size=10, bold=(ci == 2))
            cell.border    = thin_border()
            cell.alignment = Alignment(horizontal='left' if ci == 1 else 'right', vertical='center')
        if isinstance(v, (int, float)):
            ws3.cell(i, 2).number_format = '#,##0'
        ws3.row_dimensions[i].height = 22

    # Grafik uchun ma'lumotlar
    cr = 10
    for ci, h in enumerate(["Ism", "Umumiy qarz", "Qolgan"], 1):
        ws3.cell(cr, ci).value = h
        ws3.cell(cr, ci).font  = Font(bold=True, color=C_WHITE, size=10)
        ws3.cell(cr, ci).fill  = PatternFill("solid", fgColor=C_DARK)
        ws3.cell(cr, ci).alignment = Alignment(horizontal='center')

    for i, d in enumerate(active_debts[:12], 1):
        ws3.cell(cr + i, 1).value = d['name']
        ws3.cell(cr + i, 2).value = d['amount']
        ws3.cell(cr + i, 3).value = d['remaining']
        data_style(ws3, cr + i, 3, i % 2 == 0)

    n = min(len(active_debts), 12)
    if n > 0:
        # Bar chart
        ch = BarChart()
        ch.type  = "bar"
        ch.title = "Qarzdorlar holati"
        ch.style = 10
        ch.width = 22
        ch.height = 14
        dr = Reference(ws3, min_col=2, max_col=3, min_row=cr, max_row=cr + n)
        ca = Reference(ws3, min_col=1, min_row=cr + 1, max_row=cr + n)
        ch.add_data(dr, titles_from_data=True)
        ch.set_categories(ca)
        ch.series[0].graphicalProperties.solidFill = "E94560"
        ch.series[1].graphicalProperties.solidFill = "4F8EF7"
        ws3.add_chart(ch, "E2")

        # Pie chart
        pie = PieChart()
        pie.title  = "Qolgan qarz ulushi"
        pie.style  = 10
        pie.width  = 16
        pie.height = 14
        pd2 = Reference(ws3, min_col=3, min_row=cr, max_row=cr + n)
        pc  = Reference(ws3, min_col=1, min_row=cr + 1, max_row=cr + n)
        pie.add_data(pd2, titles_from_data=True)
        pie.set_categories(pc)
        ws3.add_chart(pie, "E24")

    for ci, w in enumerate([24, 18, 18], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    tmp = tempfile.NamedTemporaryFile(
        suffix=f'_excel_{chat_id}.xlsx', delete=False, dir=tempfile.gettempdir()
    )
    wb.save(tmp.name)
    return tmp.name


# ─── PDF (TUZATILGAN) ──────────────────────────────────────────────────────
def generate_pdf(active_debts, chat_id):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
    except ImportError:
        return None

    tmp = tempfile.NamedTemporaryFile(
        suffix=f'_pdf_{chat_id}.pdf', delete=False, dir=tempfile.gettempdir()
    )
    doc = SimpleDocTemplate(
        tmp.name, pagesize=A4,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=2 * cm,    bottomMargin=2 * cm
    )
    styles = getSampleStyleSheet()
    story  = []

    # ── Stillar ───────────────────────────────────────────────────────────────
    title_style = ParagraphStyle(
        'title', parent=styles['Title'],
        fontSize=20, textColor=colors.HexColor('#E94560'),
        spaceAfter=4, fontName='Helvetica-Bold'
    )
    sub_style = ParagraphStyle(
        'sub', parent=styles['Normal'],
        fontSize=10, textColor=colors.HexColor('#666666'), spaceAfter=14
    )
    hdr_style = ParagraphStyle(
        'hdr', parent=styles['Normal'],
        fontSize=13, textColor=colors.HexColor('#1A1A2E'),
        spaceBefore=14, spaceAfter=6, fontName='Helvetica-Bold'
    )

    story.append(Paragraph("QARZ HISOBCHI BOT", title_style))
    story.append(Paragraph(
        f"Hisobot: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  "
        f"Jami: {len(active_debts)} ta qarzdor",
        sub_style
    ))
    story.append(HRFlowable(width="100%", thickness=2.5,
                             color=colors.HexColor('#E94560')))
    story.append(Spacer(1, 0.4 * cm))

    # ── Umumiy ko'rsatkichlar ─────────────────────────────────────────────────
    t_amt = sum(d['amount']    for d in active_debts)
    t_rem = sum(d['remaining'] for d in active_debts)
    t_pay = t_amt - t_rem
    pct   = int(t_pay / t_amt * 100) if t_amt > 0 else 0

    story.append(Paragraph("Umumiy ko'rsatkichlar", hdr_style))
    sum_data = [
        ["Ko'rsatkich",    "Qiymat"],
        ["Jami qarzdorlar", f"{len(active_debts)} kishi"],
        ["Umumiy qarz",     f"{fmt(t_amt)} so'm"],
        ["To'langan",       f"{fmt(t_pay)} so'm"],
        ["Qolgan qarz",     f"{fmt(t_rem)} so'm"],
        ["To'lov foizi",    f"{pct}%"],
    ]
    sum_table = Table(sum_data, colWidths=[8 * cm, 8 * cm])
    sum_table.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0),  colors.HexColor('#E94560')),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1), 10),
        ('FONTSIZE',     (0, 0), (-1, 0),  11),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white]),
        ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
        ('LEFTPADDING',  (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING',   (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 8),
        ('ALIGN',        (1, 0), (1, -1),  'RIGHT'),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 0.5 * cm))

    # ── Qarzdorlar jadvali ────────────────────────────────────────────────────
    story.append(Paragraph("Qarzdorlar ro'yxati (qolgan summaga ko'ra)", hdr_style))

    table_data = [["#", "Ism", "Jami qarz", "Tolangan", "Qolgan", "%", "Valyuta"]]
    sorted_debts = sorted(active_debts, key=lambda x: x['remaining'], reverse=True)
    for i, d in enumerate(sorted_debts, 1):
        paid = d['amount'] - d['remaining']
        pct2 = int(paid / d['amount'] * 100) if d['amount'] > 0 else 0
        # PDF uchun Unicode emojilardan holi matn
        name = d['name'][:22]
        table_data.append([
            str(i), name,
            fmt(d['amount']), fmt(paid), fmt(d['remaining']),
            f"{pct2}%",
            d.get('currency', 'UZS'),
        ])

    col_w = [0.8 * cm, 5 * cm, 3.2 * cm, 3.2 * cm, 3.2 * cm, 1.4 * cm, 1.6 * cm]
    tbl = Table(table_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0),  colors.HexColor('#1A1A2E')),
        ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1), 9),
        ('FONTSIZE',     (0, 0), (-1, 0),  10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white]),
        ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
        ('LEFTPADDING',  (0, 0), (-1, -1), 6),
        ('TOPPADDING',   (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 6),
        ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN',        (1, 1), (1, -1),  'LEFT'),
        ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(tbl)

    # ── Kategoriyalar bo'yicha (agar mavjud bo'lsa) ───────────────────────────
    cats = {}
    for d in active_debts:
        c = d.get('category', '') or 'Boshqa'
        cats[c] = cats.get(c, 0) + d['remaining']
    if len(cats) > 1:
        story.append(Spacer(1, 0.4 * cm))
        story.append(Paragraph("Kategoriyalar bo'yicha", hdr_style))
        cat_data = [["Kategoriya", "Qolgan summa", "Ulush %"]]
        total_r = sum(cats.values())
        for cat, val in sorted(cats.items(), key=lambda x: x[1], reverse=True):
            pct_c = int(val / total_r * 100) if total_r > 0 else 0
            cat_data.append([cat, f"{fmt(val)} so'm", f"{pct_c}%"])
        cat_tbl = Table(cat_data, colWidths=[6 * cm, 5 * cm, 3 * cm])
        cat_tbl.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0),  colors.HexColor('#0F3460')),
            ('TEXTCOLOR',    (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',     (0, 0), (-1, 0),  'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f5f5f5'), colors.white]),
            ('GRID',         (0, 0), (-1, -1), 0.5, colors.HexColor('#dddddd')),
            ('LEFTPADDING',  (0, 0), (-1, -1), 8),
            ('TOPPADDING',   (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 7),
            ('ALIGN',        (1, 0), (-1, -1), 'CENTER'),
            ('VALIGN',       (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(cat_tbl)

    story.append(Spacer(1, 0.5 * cm))
    story.append(HRFlowable(width="100%", thickness=1,
                             color=colors.HexColor('#cccccc')))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        f"Hisobot vaqti: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        ParagraphStyle('footer', parent=styles['Normal'],
                       fontSize=8, textColor=colors.HexColor('#999999'))
    ))

    try:
        doc.build(story)
    except Exception as e:
        print(f"PDF build xatosi: {e}")
        return None

    return tmp.name
