"""
Excel/CSV shablonlarini qayta yaratish: python build_templates.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _header_row(ws):
    heads = [
        "Ism",
        "Summa",
        "Izoh",
        "Valyuta",
        "Kategoriya",
    ]
    fill = PatternFill("solid", fgColor="1A1A2E")
    font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(heads, 1):
        c = ws.cell(1, i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[1].height = 28
    widths = (22, 14, 28, 12, 14)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _borders(ws, r1, r2, c1=1, c2=5):
    thin = Side(style="thin", color="CCCCCC")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = b


def build_shablon(root: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    _header_row(ws)
    for r in range(2, 26):
        for c in range(1, 6):
            ws.cell(r, c, value=None)
    _borders(ws, 2, 25)
    ws2 = wb.create_sheet("Ko'rsatma")
    lines = [
        ("📥 Botda /import yoki Excel faylni chatga yuboring.",),
        ("",),
        ("Ustunlar (1-qator, o'zgartirmang):",),
        ("  A — Ism (yoki @username)",),
        ("  B — Summa (faqat son)",),
        ("  C — Izoh (ixtiyoriy)",),
        ("  D — Valyuta: UZS, USD yoki RUB",),
        ("  E — Kategoriya: Mahsulot, Xizmat, Shaxsiy, Boshqa (ixtiyoriy)",),
        ("",),
        ("Ma'lumot 2-qatordan boshlanadi (1-qator sarlavha).",),
    ]
    for i, row in enumerate(lines, 1):
        ws2.cell(i, 1, value=row[0])
    ws2.column_dimensions["A"].width = 72
    out = root / "qarz_import_shablon.xlsx"
    wb.save(out)
    print("OK", out.name)


def build_misol(root: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    _header_row(ws)
    examples = [
        ("Alisher Karimov", 1_500_000, "Do'kondan oylik", "UZS", "Mahsulot"),
        ("Nodira", 350_000, "", "UZS", "Xizmat"),
        ("@sardor", 500, "qarz", "USD", "Shaxsiy"),
        ("Bobur", 1200000, "Boshqa", "UZS", "Boshqa"),
    ]
    for r, row in enumerate(examples, 2):
        for c, v in enumerate(row, 1):
            ws.cell(r, c, value=v)
    _borders(ws, 2, 1 + len(examples))
    out = root / "qarz_import_misol.xlsx"
    wb.save(out)
    print("OK", out.name)


def build_csv(root: Path):
    # DictReader uchun inglizcha kalitlar (bulk_import)
    header = "name,amount,note,currency,category\n"
    body = (
        "Alisher Karimov,1500000,Do'kondan,UZS,Mahsulot\n"
        "Nodira,350000,,UZS,Xizmat\n"
    )
    p1 = root / "qarz_import_shablon.csv"
    p1.write_text("\ufeff" + header, encoding="utf-8")
    print("OK", p1.name)
    p2 = root / "qarz_import_misol.csv"
    p2.write_text("\ufeff" + header + body, encoding="utf-8")
    print("OK", p2.name)


def _style_header_cell(cell):
    fill = PatternFill("solid", fgColor="1A1A2E")
    font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="888888")
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _apply_grid(ws, r1, r2, c1, c2):
    thin = Side(style="thin", color="CCCCCC")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = b


def build_sklad_shablon(root: Path):
    """Sklad: qoldiq ro'yxati + harakatlar jurnali."""
    wb = Workbook()
    # --- Qoldiq ---
    ws = wb.active
    ws.title = "Qoldiq"
    heads = [
        "Kod",
        "Nomi",
        "Birlik",
        "Miqdor",
        "Minimal qoldiq",
        "Narh (birlik)",
        "Qiymat jami",
        "Shtellaj / joy",
        "Yetkazib beruvchi",
        "Oxirgi harakat sanasi",
        "Izoh",
    ]
    ncol = len(heads)
    for i, h in enumerate(heads, 1):
        c = ws.cell(1, i, value=h)
        _style_header_cell(c)
    ws.row_dimensions[1].height = 28
    widths = (12, 28, 10, 12, 14, 14, 14, 16, 20, 16, 26)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    last_data_row = 200
    for r in range(2, last_data_row + 1):
        ws.cell(r, 7, value=f'=IF(OR(D{r}="",F{r}=""),"",D{r}*F{r})')
        for c in (4, 5, 6):
            ws.cell(r, c).alignment = Alignment(horizontal="right", vertical="center")
    _apply_grid(ws, 2, last_data_row, 1, ncol)

    # --- Harakatlar ---
    w2 = wb.create_sheet("Harakatlar")
    h2 = [
        "Sana",
        "Tur",
        "Kod",
        "Nomi",
        "Miqdor",
        "Birlik",
        "Hujjat / invoice",
        "Izoh",
    ]
    for i, h in enumerate(h2, 1):
        _style_header_cell(w2.cell(1, i, value=h))
    w2.row_dimensions[1].height = 28
    w2_widths = (14, 12, 12, 26, 12, 10, 22, 30)
    for i, w in enumerate(w2_widths, 1):
        w2.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, 501):
        w2.cell(r, 2).alignment = Alignment(horizontal="center", vertical="center")
    _apply_grid(w2, 2, 500, 1, len(h2))

    # --- Ko'rsatma ---
    w3 = wb.create_sheet("Ko'rsatma")
    lines = [
        "SKLAD SHABLONI",
        "",
        "Varaq «Qoldiq» — asosiy ro'yxat: har bir qator bitta mahsulot.",
        "  • G ustuni (Qiymat jami) formulali: Miqdor × Narh (o'zgartirmang).",
        "  • Minimal qoldiqdan past bo'lsa, alohida filtr yoki rang bilan ajrating.",
        "",
        "Varaq «Harakatlar» — kirim/chiqim yozuvlari (sana, tur, miqdor).",
        "  • Tur: Kirim yoki Chiqim (yoki O'tkazma).",
        "",
        "Masalan uchun: sklad_misol.xlsx faylini oching.",
    ]
    for i, line in enumerate(lines, 1):
        w3.cell(i, 1, value=line)
    w3.column_dimensions["A"].width = 78

    out = root / "sklad_shablon.xlsx"
    wb.save(out)
    print("OK", out.name)


def build_sklad_misol(root: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Qoldiq"
    heads = [
        "Kod",
        "Nomi",
        "Birlik",
        "Miqdor",
        "Minimal qoldiq",
        "Narh (birlik)",
        "Qiymat jami",
        "Shtellaj / joy",
        "Yetkazib beruvchi",
        "Oxirgi harakat sanasi",
        "Izoh",
    ]
    ncol = len(heads)
    for i, h in enumerate(heads, 1):
        _style_header_cell(ws.cell(1, i, value=h))
    ws.row_dimensions[1].height = 28
    widths = (12, 28, 10, 12, 14, 14, 14, 16, 20, 16, 26)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    rows = [
        ("SKU-001", "Sement M-400", "qop", 120, 20, 95000, None, "A-1", "OOO Build", "2026-04-01", ""),
        ("SKU-002", "Armatura 12mm", "tonna", 3.5, 1, 12500000, None, "B-3", "Metal Trade", "2026-04-05", "rezerva"),
        ("SKU-003", "Gipsokarton 12.5mm", "varaq", 80, 15, 45000, None, "A-2", "StroyMart", "2026-04-08", ""),
    ]
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            if c == 7:
                ws.cell(r, 7, value=f'=IF(OR(D{r}="",F{r}=""),"",D{r}*F{r})')
            else:
                ws.cell(r, c, value=v)
    _apply_grid(ws, 2, 1 + len(rows), 1, ncol)
    for rr in range(2, 2 + len(rows)):
        for c in (4, 5, 6):
            ws.cell(rr, c).alignment = Alignment(horizontal="right", vertical="center")

    w2 = wb.create_sheet("Harakatlar")
    h2 = ["Sana", "Tur", "Kod", "Nomi", "Miqdor", "Birlik", "Hujjat / invoice", "Izoh"]
    for i, h in enumerate(h2, 1):
        _style_header_cell(w2.cell(1, i, value=h))
    w2.row_dimensions[1].height = 28
    w2_widths = (14, 12, 12, 26, 12, 10, 22, 30)
    for i, w in enumerate(w2_widths, 1):
        w2.column_dimensions[get_column_letter(i)].width = w
    mov = [
        ("2026-04-10", "Kirim", "SKU-001", "Sement M-400", 40, "qop", "INV-104", ""),
        ("2026-04-09", "Chiqim", "SKU-002", "Armatura 12mm", 0.5, "tonna", "Nakladnaya-22", "ob'ekt"),
    ]
    for r, row in enumerate(mov, 2):
        for c, v in enumerate(row, 1):
            w2.cell(r, c, value=v)
    _apply_grid(w2, 2, 1 + len(mov), 1, len(h2))

    out = root / "sklad_misol.xlsx"
    wb.save(out)
    print("OK", out.name)


if __name__ == "__main__":
    root = Path(__file__).resolve().parent
    build_shablon(root)
    build_misol(root)
    build_csv(root)
    build_sklad_shablon(root)
    build_sklad_misol(root)
